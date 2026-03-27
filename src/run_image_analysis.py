"""
Image analysis batch runner — re-scrapes UNCERTAIN rows for product images,
compares them visually, and upgrades confident matches to LIKELY in the Excel file.

Pipeline:
  1. Local CV (instant, free, unlimited):
       perceptual hash + ORB feature matching + histogram
  2. Gemini Vision (free, 1500/day, 15 RPM):
       only for MEDIUM local CV results (0.3-0.7 similarity)
  3. Excel update:
       HIGH similarity -> upgrade UNCERTAIN -> LIKELY (confidence +15)
       Gemini YES/LIKELY + confidence >= 70 -> upgrade UNCERTAIN -> LIKELY
       Otherwise -> leave UNCERTAIN (visual alone not enough to mark NO)

Usage:
    uv run python src/run_image_analysis.py Anchor
    uv run python src/run_image_analysis.py Dorman
    uv run python src/run_image_analysis.py Anchor --limit 100
    uv run python src/run_image_analysis.py Dorman --no-gemini   # local CV only
    uv run python src/run_image_analysis.py --count              # just print stats, no changes
"""

import sys
import os
import re
import json
import argparse
import time
from pathlib import Path
from datetime import datetime

# Allow running from project root or src/
_src = Path(__file__).parent
if str(_src) not in sys.path:
    sys.path.insert(0, str(_src))
if str(_src.parent) not in sys.path:
    sys.path.insert(0, str(_src.parent))

import openpyxl
from openpyxl.styles import PatternFill

EXCEL_FILE = Path(__file__).parent.parent / "FISHER SKP INTERCHANGE 20260302.xlsx"
PROGRESS_DIR = Path(__file__).parent.parent / "output"

FILL_YES       = PatternFill("solid", fgColor="00FF00")
FILL_LIKELY    = PatternFill("solid", fgColor="90EE90")
FILL_UNCERTAIN = PatternFill("solid", fgColor="FFFF00")
FILL_NO        = PatternFill("solid", fgColor="FF0000")

FILL_MAP = {
    "YES":       FILL_YES,
    "LIKELY":    FILL_LIKELY,
    "UNCERTAIN": FILL_UNCERTAIN,
    "NO":        FILL_NO,
}


# ─── helpers ─────────────────────────────────────────────────────────────────

def _save_workbook(wb, path: Path):
    """Crash-safe save: write to .tmp then replace."""
    tmp = path.with_suffix(".tmp")
    wb.save(str(tmp))
    os.replace(str(tmp), str(path))


def _skip_value(v) -> bool:
    if v is None:
        return True
    s = str(v).strip().lower()
    return s in ("", "n/a", "-", "0", "none", "tbd", "?")


# ─── local CV similarity (image_compare.py logic inlined for portability) ────

def _local_cv_similarity(url1: str, url2: str) -> dict:
    """
    Compare two image URLs using local CV (perceptual hash + ORB + histogram).
    Returns {"score": float 0-1, "detail": {...}, "error": str|None}
    """
    result = {"score": 0.0, "detail": {}, "error": None}
    try:
        import requests as req
        from io import BytesIO
        from PIL import Image
        import imagehash
        import cv2
        import numpy as np

        headers = {"User-Agent": "Mozilla/5.0 Chrome/120.0.0.0"}

        r1 = req.get(url1, timeout=12, headers=headers)
        r2 = req.get(url2, timeout=12, headers=headers)
        r1.raise_for_status()
        r2.raise_for_status()

        # Perceptual hash
        pil1 = Image.open(BytesIO(r1.content)).convert("RGB")
        pil2 = Image.open(BytesIO(r2.content)).convert("RGB")
        h1, h2 = imagehash.phash(pil1), imagehash.phash(pil2)
        phash_score = max(0.0, (64 - (h1 - h2)) / 64)
        result["detail"]["phash"] = round(phash_score, 3)

        # OpenCV methods
        def pil_to_cv(img):
            import numpy as np
            return cv2.cvtColor(np.array(img), cv2.COLOR_RGB2BGR)

        cv1, cv2_ = pil_to_cv(pil1), pil_to_cv(pil2)

        # ORB feature matching
        orb = cv2.ORB_create(nfeatures=500)
        g1 = cv2.cvtColor(cv1, cv2.COLOR_BGR2GRAY)
        g2 = cv2.cvtColor(cv2_, cv2.COLOR_BGR2GRAY)
        kp1, d1 = orb.detectAndCompute(g1, None)
        kp2, d2 = orb.detectAndCompute(g2, None)
        if d1 is not None and d2 is not None and len(d1) >= 10 and len(d2) >= 10:
            bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)
            matches = bf.match(d1, d2)
            good = [m for m in matches if m.distance < 50]
            orb_score = min(len(good) / max(min(len(kp1), len(kp2)), 1), 1.0)
        else:
            orb_score = 0.0
        result["detail"]["orb"] = round(orb_score, 3)

        # Histogram
        scores = []
        for ch in range(3):
            h_a = cv2.calcHist([cv1], [ch], None, [256], [0, 256])
            h_b = cv2.calcHist([cv2_], [ch], None, [256], [0, 256])
            scores.append(cv2.compareHist(h_a, h_b, cv2.HISTCMP_CORREL))
        hist_score = max(0.0, sum(scores) / 3)
        result["detail"]["hist"] = round(hist_score, 3)

        # Weighted combination
        combined = (phash_score * 0.4 + orb_score * 0.4 + hist_score * 0.2)
        result["score"] = round(combined, 3)

    except Exception as e:
        result["error"] = str(e)
        print(f"    [LocalCV] Error: {e}")

    return result


# ─── main analysis ────────────────────────────────────────────────────────────

def _get_vision_backend(use_local: bool, use_gemini: bool):
    """
    Return (compare_fn, label) for the selected vision backend.
    Priority: local (Ollama) > Gemini > None
    """
    if use_local:
        from local_vision import compare_images as local_compare, check_ollama
        ok, msg = check_ollama()
        if ok:
            print("[Vision] Using LOCAL Ollama vision (free, unlimited)")
            return local_compare, "local"
        else:
            print(f"[Vision] Ollama not available: {msg}")
            print("[Vision] Falling back to Gemini Vision...")

    if use_gemini:
        from gemini_vision import compare_images as gemini_compare, is_available
        if is_available():
            print("[Vision] Using Gemini Vision (free API, quota-limited)")
            return gemini_compare, "gemini"
        else:
            print("[Vision] Gemini quota exhausted for today.")

    print("[Vision] No vision backend available — local CV only")
    return None, "none"


def _scrape_with_retry(part_number: str, brand: str, max_attempts: int = 2) -> dict:
    """
    Scrape RockAuto with retry using subprocess isolation.
    Returns the scrape result dict; result["found"] will be False on all failures.
    """
    from scraper_subprocess import scrape_rockauto_subprocess as scrape_rockauto
    for attempt in range(max_attempts):
        try:
            result = scrape_rockauto(part_number, brand=brand)
            err = result.get("error") or ""
            if any(kw in str(err).lower() for kw in
                   ("playwright", "browsertype", "context", "timeout", "profile")):
                raise RuntimeError(f"Browser error: {err}")
            return result
        except Exception as e:
            if attempt < max_attempts - 1:
                print(f"    [Retry {attempt+1}] Scrape error: {e} — retrying...")
                time.sleep(3)
            else:
                return {"found": False, "image_url": None, "error": str(e)}
    return {"found": False, "image_url": None, "error": "max retries exceeded"}


def run_analysis(sheet_name: str, limit: int = 0, use_gemini: bool = True,
                 use_local: bool = False, count_only: bool = False):
    if not EXCEL_FILE.exists():
        print(f"ERROR: Excel file not found: {EXCEL_FILE}")
        sys.exit(1)

    print(f"\n{'=' * 60}")
    print(f"  Image Analysis — {sheet_name}")
    print(f"{'=' * 60}")

    # Load workbook
    print("Loading workbook...")
    wb = openpyxl.load_workbook(str(EXCEL_FILE))
    ws = wb[sheet_name]

    brand_filter = sheet_name.upper()
    if brand_filter == "FOUR SEASONS":
        brand_filter = "FOUR SEASONS"

    # Collect UNCERTAIN rows
    uncertain_rows = []
    for row_num in range(2, ws.max_row + 1):
        supplier = ws.cell(row=row_num, column=2).value
        match_result = ws.cell(row=row_num, column=10).value  # col J
        part_num = ws.cell(row=row_num, column=3).value       # col C
        skp_num = ws.cell(row=row_num, column=6).value        # col F
        part_type = ws.cell(row=row_num, column=1).value      # col A

        if str(supplier or "").strip().upper() != brand_filter:
            continue
        if match_result != "UNCERTAIN":
            continue
        if _skip_value(part_num) or _skip_value(skp_num):
            continue

        uncertain_rows.append({
            "row_num": row_num,
            "part_num": str(part_num).strip(),
            "skp_num": str(skp_num).strip(),
            "part_type": str(part_type or "").strip(),
        })

    total_uncertain = len(uncertain_rows)
    print(f"Found {total_uncertain:,} UNCERTAIN rows in '{sheet_name}'")

    if count_only:
        # Just print current stats and exit
        _print_stats(ws, sheet_name)
        wb.close()
        return

    if total_uncertain == 0:
        print("Nothing to analyze.")
        wb.close()
        return

    # Apply limit
    batch = uncertain_rows[:limit] if limit > 0 else uncertain_rows
    print(f"Processing {len(batch):,} rows" + (f" (limit={limit})" if limit else ""))

    # Select vision backend
    vision_fn, vision_label = _get_vision_backend(use_local=use_local, use_gemini=use_gemini)
    print()

    # Stats tracking
    upgraded = 0
    no_images = 0
    high_cv = 0
    medium_cv = 0
    low_cv = 0
    gemini_used = 0
    errors = 0
    consecutive_errors = 0
    MAX_CONSECUTIVE_ERRORS = 10  # abort if browser is completely stuck

    for i, row_data in enumerate(batch):
        row_num = row_data["row_num"]
        part_num = row_data["part_num"]
        skp_num = row_data["skp_num"]
        part_type = row_data["part_type"]

        print(f"[{i+1}/{len(batch)}] Row {row_num}: {brand_filter} {part_num} vs SKP {skp_num}")

        try:
            # ── Step 1: Scrape both parts for image URLs ──────────────────
            brand_data = _scrape_with_retry(part_num, brand=brand_filter)
            skp_data   = _scrape_with_retry(skp_num,  brand="SKP")

            img1 = brand_data.get("image_url")
            img2 = skp_data.get("image_url")

            if not img1 or not img2:
                missing = []
                if not img1: missing.append(f"{brand_filter} {part_num}")
                if not img2: missing.append(f"SKP {skp_num}")
                print(f"    No image for: {', '.join(missing)} — skipping")
                no_images += 1
                continue

            print(f"    Images: {img1}")
            print(f"            {img2}")

            # ── Step 2: Local CV ──────────────────────────────────────────
            cv_result = _local_cv_similarity(img1, img2)
            cv_score = cv_result["score"]
            cv_detail = cv_result["detail"]
            print(f"    Local CV: {cv_score:.3f}  {cv_detail}")

            # Classify CV result
            final_match = None
            final_confidence = None
            final_reason = None

            if cv_result.get("error"):
                errors += 1
                continue

            if cv_score >= 0.75:
                # HIGH — upgrade immediately, no AI vision needed
                high_cv += 1
                final_match = "LIKELY"
                final_confidence = 72
                final_reason = f"Image comparison: HIGH visual similarity (CV={cv_score:.2f}). Parts appear visually identical or near-identical."

            elif cv_score >= 0.35 and vision_fn is not None:
                # MEDIUM — use AI vision backend for confirmation
                medium_cv += 1
                print(f"    [{vision_label.upper()} Vision] comparing...")
                gv = vision_fn(img1, img2, part_type=part_type,
                               brand1=brand_filter, brand2="SKP")
                gemini_used += 1
                err = gv.get("error")
                if err and "quota" in str(err).lower():
                    print(f"    Vision quota exhausted — skipping for remaining MEDIUM cases")
                    vision_fn = None  # disable for rest of run
                elif not err:
                    gm = gv["match"]
                    gc = gv["confidence"]
                    gr = gv["reasoning"]
                    print(f"    [{vision_label.upper()}] {gm} ({gc}%) — {gr[:80]}")
                    if gm in ("YES", "LIKELY") and gc >= 65:
                        final_match = "LIKELY"
                        final_confidence = min(gc, 80)
                        final_reason = f"[{vision_label}] {gm} ({gc}%). {gr}"
                    else:
                        print(f"    Not upgrading ({vision_label}: {gm}, confidence {gc}%)")
                else:
                    print(f"    Vision error: {err}")

            else:
                # LOW cv_score — don't upgrade, don't downgrade
                low_cv += 1
                print(f"    LOW visual similarity ({cv_score:.3f}) — keeping UNCERTAIN")

            # ── Step 3: Write upgrade to Excel ────────────────────────────
            if final_match == "LIKELY":
                upgraded += 1
                ts = datetime.now().strftime("%Y-%m-%d %H:%M")
                old_reason = ws.cell(row=row_num, column=12).value or ""
                new_reason = f"[Image] {final_reason[:200]}"

                ws.cell(row=row_num, column=10).value = "LIKELY"
                ws.cell(row=row_num, column=10).fill = FILL_LIKELY
                ws.cell(row=row_num, column=11).value = final_confidence
                ws.cell(row=row_num, column=12).value = new_reason
                ws.cell(row=row_num, column=16).value = ts  # col P = LAST CHECKED

                print(f"    -> UPGRADED to LIKELY ({final_confidence}%)")

                # Crash-safe save after every upgrade
                _save_workbook(wb, EXCEL_FILE)

        except KeyboardInterrupt:
            print("\n\nStopped by user.")
            break
        except Exception as e:
            errors += 1
            consecutive_errors += 1
            print(f"    ERROR: {e}")
            if consecutive_errors >= MAX_CONSECUTIVE_ERRORS:
                print(f"\nABORTING: {MAX_CONSECUTIVE_ERRORS} consecutive errors — browser may be stuck.")
                print("Try restarting Chrome manually, then re-run.")
                break
            continue

        # Reset consecutive error counter on any successful row
        consecutive_errors = 0

    # ── Final report ───────────────────────────────────────────────────────
    print(f"\n{'=' * 60}")
    print(f"  COMPLETED — {sheet_name} Image Analysis")
    print(f"{'=' * 60}")
    print(f"  Rows processed:        {len(batch):>6,}")
    print(f"  No image available:    {no_images:>6,}")
    print(f"  HIGH CV (>=0.75):      {high_cv:>6,}")
    print(f"  MEDIUM CV (0.35-0.75): {medium_cv:>6,}")
    print(f"  LOW CV (<0.35):        {low_cv:>6,}")
    print(f"  Gemini Vision calls:   {gemini_used:>6,}")
    print(f"  Errors:                {errors:>6,}")
    print(f"  --------------------------------------")
    print(f"  UPGRADED to LIKELY:    {upgraded:>6,}")
    print()
    _print_stats(ws, sheet_name)

    wb.close()


def _print_stats(ws, sheet_name: str):
    """Print current YES/LIKELY/UNCERTAIN/NO counts for the sheet."""
    counts = {"YES": 0, "LIKELY": 0, "UNCERTAIN": 0, "NO": 0, "BLANK": 0}
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = row[9] if len(row) > 9 else None
        if v in counts:
            counts[v] += 1
        else:
            counts["BLANK"] += 1

    processed = counts["YES"] + counts["LIKELY"] + counts["UNCERTAIN"] + counts["NO"]
    pct = lambda n: f"{n / processed * 100:.1f}%" if processed > 0 else "N/A"
    confirmed = counts["YES"] + counts["LIKELY"]

    print(f"  Current counts for '{sheet_name}':")
    print(f"    YES:       {counts['YES']:>6,}  ({pct(counts['YES'])})")
    print(f"    LIKELY:    {counts['LIKELY']:>6,}  ({pct(counts['LIKELY'])})")
    print(f"    UNCERTAIN: {counts['UNCERTAIN']:>6,}  ({pct(counts['UNCERTAIN'])})")
    print(f"    NO:        {counts['NO']:>6,}  ({pct(counts['NO'])})")
    print(f"    ---------------------------------")
    print(f"    Confirmed: {confirmed:>6,}  ({pct(confirmed)})  [YES + LIKELY]")
    print(f"    Pending:   {counts['UNCERTAIN']:>6,}  ({pct(counts['UNCERTAIN'])})  [UNCERTAIN]")


def main():
    parser = argparse.ArgumentParser(
        description="Run image comparison on UNCERTAIN rows to upgrade matches"
    )
    parser.add_argument("sheet", nargs="?", choices=["Anchor", "Dorman", "GMB", "SMP", "Four Seasons "],
                        help="Sheet to process")
    parser.add_argument("--limit", type=int, default=0,
                        help="Max rows to process (0 = all, default 0)")
    parser.add_argument("--local", action="store_true",
                        help="Use local Ollama vision model (free, unlimited, GPU-accelerated)")
    parser.add_argument("--no-gemini", action="store_true",
                        help="Skip Gemini Vision (use local CV or --local only)")
    parser.add_argument("--model", default=None,
                        help="Ollama model name (default: llava-phi3). Use with --local.")
    parser.add_argument("--count", action="store_true",
                        help="Just print current stats, no processing")

    args = parser.parse_args()

    if not args.sheet and not args.count:
        parser.print_help()
        print("\nExamples:")
        print("  uv run python src/run_image_analysis.py Dorman --local        # local Ollama (recommended)")
        print("  uv run python src/run_image_analysis.py Dorman --limit 50     # Gemini, 50 rows")
        print("  uv run python src/run_image_analysis.py --count               # stats only")
        sys.exit(1)

    # Override model if specified
    if args.model and args.local:
        import local_vision
        local_vision.DEFAULT_MODEL = args.model

    if args.count:
        import openpyxl as xl
        wb = xl.load_workbook(str(EXCEL_FILE), read_only=True, data_only=True)
        for sn in ["Anchor", "Dorman", "GMB", "SMP", "Four Seasons "]:
            if sn in wb.sheetnames:
                ws = wb[sn]
                print(f"\n--- {sn} ---")
                _print_stats(ws, sn)
        wb.close()
        return

    run_analysis(
        sheet_name=args.sheet,
        limit=args.limit,
        use_local=args.local,
        use_gemini=not args.no_gemini,
        count_only=False,
    )


if __name__ == "__main__":
    main()
