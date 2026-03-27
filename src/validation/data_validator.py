"""
Data Validator
Data Quality & Validation Layer (Priority 4)

Comprehensive input data validation system for:
- Excel data structure and format validation
- Part number format validation
- OEM reference validation
- Brand and category validation
- Data completeness assessment
"""

import re
import pandas as pd
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import sys

# Add parent directory for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

class DataValidator:
    """Comprehensive data validation system for Parts Agent."""

    def __init__(self, excel_path: str = None):
        """Initialize the data validator."""
        if excel_path is None:
            excel_path = str(Path(__file__).parent.parent.parent / "FISHER SKP INTERCHANGE 20260302.xlsx")

        self.excel_path = excel_path

        # Expected sheet structure
        self.required_sheets = ["GMB", "Four Seasons ", "SMP", "Anchor", "Dorman"]

        # Expected columns
        self.required_columns = {
            'A': 'PART TYPE',
            'B': 'CURRENT SUPPLIER',
            'C': 'PART #',
            'D': 'CALL12',
            'E': 'DLS',
            'F': 'SKP PART #',
            'G': 'SKP PART # - Check',
            'H': 'SKP QUOTE',
            'I': 'Notes'
        }

        # Output columns (written by the system)
        self.output_columns = {
            'J': 'MATCH RESULT',
            'K': 'CONFIDENCE %',
            'L': 'MATCH REASON',
            'M': 'FITMENT MATCH',
            'N': 'DESC MATCH',
            'O': 'MISSING INFO',
            'P': 'LAST CHECKED'
        }

        # Valid values for output columns
        self.valid_match_results = {'YES', 'LIKELY', 'UNCERTAIN', 'NO'}
        self.valid_fitment_values = {'YES', 'NO', 'UNKNOWN'}
        self.valid_desc_match_values = {'YES', 'NO', 'PARTIAL'}

        # Part number validation patterns
        self.part_number_patterns = {
            'alphanumeric': re.compile(r'^[A-Za-z0-9\-\_]+$'),
            'has_content': re.compile(r'[A-Za-z0-9]'),
            'reasonable_length': lambda x: 2 <= len(str(x).strip()) <= 50
        }

        # OEM reference patterns
        self.oem_patterns = {
            'alphanumeric_extended': re.compile(r'^[A-Za-z0-9\-\_\.\+\s]+$'),
            'minimum_length': lambda x: len(str(x).strip()) >= 3
        }

        # Known brands
        self.valid_brands = {
            'ANCHOR', 'DORMAN', 'GMB', 'SMP', 'FOUR SEASONS', 'SKP'
        }

        # Common part type categories
        self.common_part_types = {
            'ENGINE MOUNT', 'BRAKE PAD', 'FILTER', 'SENSOR', 'BELT',
            'GASKET', 'BEARING', 'SEAL', 'PUMP', 'VALVE', 'HOSE'
        }

        # Validation statistics
        self.validation_stats = {
            'total_rows_validated': 0,
            'validation_errors': 0,
            'validation_warnings': 0,
            'last_validation': None
        }

        print("[DATA_VALIDATOR] Initialized with comprehensive validation rules")

    def validate_excel_structure(self) -> Dict:
        """Validate the basic Excel file structure."""
        validation_result = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'details': {}
        }

        try:
            # Check file exists
            if not Path(self.excel_path).exists():
                validation_result['valid'] = False
                validation_result['errors'].append(f"Excel file not found: {self.excel_path}")
                return validation_result

            # Load Excel file and check sheets
            import openpyxl
            workbook = openpyxl.load_workbook(self.excel_path)

            # Check required sheets
            available_sheets = workbook.sheetnames
            missing_sheets = [sheet for sheet in self.required_sheets if sheet not in available_sheets]
            extra_sheets = [sheet for sheet in available_sheets if sheet not in self.required_sheets]

            if missing_sheets:
                validation_result['valid'] = False
                validation_result['errors'].append(f"Missing required sheets: {missing_sheets}")

            if extra_sheets:
                validation_result['warnings'].append(f"Unexpected sheets found: {extra_sheets}")

            # Check column structure for each sheet
            sheet_details = {}
            for sheet_name in self.required_sheets:
                if sheet_name in available_sheets:
                    sheet = workbook[sheet_name]
                    sheet_validation = self._validate_sheet_structure(sheet, sheet_name)
                    sheet_details[sheet_name] = sheet_validation

                    if not sheet_validation['valid']:
                        validation_result['valid'] = False
                        validation_result['errors'].extend(sheet_validation['errors'])

                    validation_result['warnings'].extend(sheet_validation['warnings'])

            validation_result['details']['sheets'] = sheet_details
            validation_result['details']['total_sheets'] = len(available_sheets)
            validation_result['details']['valid_sheets'] = len([s for s in sheet_details.values() if s.get('valid', False)])

            workbook.close()

        except Exception as e:
            validation_result['valid'] = False
            validation_result['errors'].append(f"Error reading Excel file: {str(e)}")

        return validation_result

    def _validate_sheet_structure(self, sheet, sheet_name: str) -> Dict:
        """Validate individual sheet structure."""
        validation_result = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'row_count': 0,
            'column_validation': {}
        }

        try:
            # Check header row (row 1)
            for col_letter, expected_name in self.required_columns.items():
                cell_value = sheet[f'{col_letter}1'].value
                if cell_value != expected_name:
                    validation_result['warnings'].append(
                        f"Sheet {sheet_name} column {col_letter}: Expected '{expected_name}', found '{cell_value}'"
                    )

            # Count non-empty rows
            row_count = 0
            for row in sheet.iter_rows(min_row=2):  # Skip header
                if any(cell.value for cell in row[:9]):  # Check first 9 columns
                    row_count += 1

            validation_result['row_count'] = row_count

            if row_count == 0:
                validation_result['warnings'].append(f"Sheet {sheet_name} appears to be empty")

        except Exception as e:
            validation_result['valid'] = False
            validation_result['errors'].append(f"Error validating sheet {sheet_name}: {str(e)}")

        return validation_result

    def validate_row_data(self, row_data: Dict) -> Dict:
        """Validate a single row of data."""
        validation_result = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'quality_score': 1.0,
            'validation_details': {}
        }

        try:
            # Validate part type
            part_type_validation = self._validate_part_type(row_data.get('part_type'))
            validation_result['validation_details']['part_type'] = part_type_validation
            if not part_type_validation['valid']:
                validation_result['valid'] = False
                validation_result['errors'].extend(part_type_validation['errors'])
            validation_result['warnings'].extend(part_type_validation['warnings'])

            # Validate brand/supplier
            brand_validation = self._validate_brand(row_data.get('current_supplier'))
            validation_result['validation_details']['brand'] = brand_validation
            if not brand_validation['valid']:
                validation_result['valid'] = False
                validation_result['errors'].extend(brand_validation['errors'])
            validation_result['warnings'].extend(brand_validation['warnings'])

            # Validate part numbers
            part_number_validation = self._validate_part_number(row_data.get('part_number'))
            validation_result['validation_details']['part_number'] = part_number_validation
            if not part_number_validation['valid']:
                validation_result['valid'] = False
                validation_result['errors'].extend(part_number_validation['errors'])
            validation_result['warnings'].extend(part_number_validation['warnings'])

            skp_number_validation = self._validate_part_number(row_data.get('skp_part_number'))
            validation_result['validation_details']['skp_part_number'] = skp_number_validation
            if not skp_number_validation['valid']:
                validation_result['valid'] = False
                validation_result['errors'].extend(skp_number_validation['errors'])
            validation_result['warnings'].extend(skp_number_validation['warnings'])

            # Validate sales data if present
            if 'call12' in row_data:
                sales_validation = self._validate_sales_data(row_data['call12'])
                validation_result['validation_details']['sales'] = sales_validation
                validation_result['warnings'].extend(sales_validation['warnings'])

            # Calculate quality score
            validation_result['quality_score'] = self._calculate_data_quality_score(
                validation_result['validation_details']
            )

            self.validation_stats['total_rows_validated'] += 1
            if not validation_result['valid']:
                self.validation_stats['validation_errors'] += 1
            if validation_result['warnings']:
                self.validation_stats['validation_warnings'] += 1

        except Exception as e:
            validation_result['valid'] = False
            validation_result['errors'].append(f"Error validating row data: {str(e)}")

        return validation_result

    def _validate_part_type(self, part_type: Any) -> Dict:
        """Validate part type field."""
        result = {'valid': True, 'errors': [], 'warnings': []}

        if not part_type:
            result['valid'] = False
            result['errors'].append("Part type is required")
            return result

        part_type_str = str(part_type).strip().upper()

        if len(part_type_str) < 3:
            result['warnings'].append("Part type is very short")

        if not any(keyword in part_type_str for keyword in ['ENGINE', 'BRAKE', 'FILTER', 'SENSOR', 'BELT', 'MOUNT', 'PAD']):
            result['warnings'].append("Part type doesn't match common patterns")

        return result

    def _validate_brand(self, brand: Any) -> Dict:
        """Validate brand/supplier field."""
        result = {'valid': True, 'errors': [], 'warnings': []}

        if not brand:
            result['valid'] = False
            result['errors'].append("Brand/supplier is required")
            return result

        brand_str = str(brand).strip().upper()

        if brand_str not in self.valid_brands:
            result['warnings'].append(f"Brand '{brand_str}' is not in known brands list")

        return result

    def _validate_part_number(self, part_number: Any) -> Dict:
        """Validate part number format."""
        result = {'valid': True, 'errors': [], 'warnings': []}

        if not part_number:
            result['valid'] = False
            result['errors'].append("Part number is required")
            return result

        part_number_str = str(part_number).strip()

        # Check for blank equivalents
        blank_equivalents = {'n/a', '-', '0', 'none', 'tbd', '?', ''}
        if part_number_str.lower() in blank_equivalents:
            result['valid'] = False
            result['errors'].append("Part number appears to be blank or placeholder")
            return result

        # Check length
        if not self.part_number_patterns['reasonable_length'](part_number_str):
            result['warnings'].append(f"Part number length ({len(part_number_str)}) is unusual")

        # Check format
        if not self.part_number_patterns['alphanumeric'].match(part_number_str):
            result['warnings'].append("Part number contains unusual characters")

        if not self.part_number_patterns['has_content'].search(part_number_str):
            result['valid'] = False
            result['errors'].append("Part number has no alphanumeric content")

        return result

    def _validate_sales_data(self, sales_data: Any) -> Dict:
        """Validate sales/call12 data."""
        result = {'valid': True, 'errors': [], 'warnings': []}

        try:
            if sales_data is not None:
                sales_value = float(sales_data)
                if sales_value < 0:
                    result['warnings'].append("Negative sales value")
                elif sales_value > 10000:
                    result['warnings'].append("Unusually high sales value")
        except (ValueError, TypeError):
            result['warnings'].append("Sales data is not numeric")

        return result

    def _calculate_data_quality_score(self, validation_details: Dict) -> float:
        """Calculate overall data quality score (0-1)."""
        try:
            total_fields = len(validation_details)
            if total_fields == 0:
                return 0.0

            quality_scores = []

            for field_name, field_validation in validation_details.items():
                field_score = 1.0

                # Penalize errors heavily
                if not field_validation.get('valid', True):
                    field_score = 0.0
                # Penalize warnings lightly
                elif field_validation.get('warnings'):
                    field_score = 0.8

                quality_scores.append(field_score)

            return sum(quality_scores) / len(quality_scores)

        except Exception as e:
            print(f"Warning: Error calculating quality score: {e}")
            return 0.5

    def validate_output_data(self, output_data: Dict) -> Dict:
        """Validate processing output data."""
        validation_result = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'quality_score': 1.0
        }

        try:
            # Validate match result
            match_result = output_data.get('match_result')
            if match_result not in self.valid_match_results:
                validation_result['valid'] = False
                validation_result['errors'].append(f"Invalid match result: {match_result}")

            # Validate confidence
            confidence = output_data.get('confidence')
            if confidence is not None:
                try:
                    conf_value = int(confidence)
                    if not (0 <= conf_value <= 100):
                        validation_result['valid'] = False
                        validation_result['errors'].append(f"Confidence must be 0-100, got: {conf_value}")
                except (ValueError, TypeError):
                    validation_result['valid'] = False
                    validation_result['errors'].append("Confidence must be an integer")

            # Validate fitment match
            fitment = output_data.get('fitment_match')
            if fitment and fitment not in self.valid_fitment_values:
                validation_result['warnings'].append(f"Unusual fitment value: {fitment}")

            # Validate description match
            desc_match = output_data.get('desc_match')
            if desc_match and desc_match not in self.valid_desc_match_values:
                validation_result['warnings'].append(f"Unusual description match value: {desc_match}")

            # Validate match reason exists
            if not output_data.get('match_reason'):
                validation_result['warnings'].append("Missing match reason explanation")

        except Exception as e:
            validation_result['valid'] = False
            validation_result['errors'].append(f"Error validating output data: {str(e)}")

        return validation_result

    def validate_batch_consistency(self, batch_data: List[Dict]) -> Dict:
        """Validate consistency across a batch of data."""
        validation_result = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'consistency_score': 1.0,
            'batch_statistics': {}
        }

        try:
            if not batch_data:
                validation_result['valid'] = False
                validation_result['errors'].append("Empty batch data")
                return validation_result

            # Check brand consistency
            brands = [item.get('current_supplier') for item in batch_data if item.get('current_supplier')]
            unique_brands = set(brands)

            if len(unique_brands) > 1:
                validation_result['warnings'].append(f"Multiple brands in batch: {unique_brands}")

            # Check part type distribution
            part_types = [item.get('part_type') for item in batch_data if item.get('part_type')]
            type_counts = {}
            for pt in part_types:
                type_counts[pt] = type_counts.get(pt, 0) + 1

            # Statistical analysis
            validation_result['batch_statistics'] = {
                'total_items': len(batch_data),
                'unique_brands': len(unique_brands),
                'brand_distribution': dict(type_counts) if len(brands) == len(batch_data) else {},
                'part_type_distribution': type_counts,
                'missing_data_count': sum(1 for item in batch_data if not item.get('part_number') or not item.get('skp_part_number'))
            }

            # Calculate consistency score
            missing_ratio = validation_result['batch_statistics']['missing_data_count'] / len(batch_data)
            validation_result['consistency_score'] = max(0.0, 1.0 - missing_ratio)

        except Exception as e:
            validation_result['valid'] = False
            validation_result['errors'].append(f"Error validating batch consistency: {str(e)}")

        return validation_result

    def get_validation_report(self) -> Dict:
        """Generate comprehensive validation report."""
        try:
            # Validate Excel structure
            structure_validation = self.validate_excel_structure()

            report = {
                'timestamp': datetime.now().isoformat(),
                'excel_structure': structure_validation,
                'validation_statistics': self.validation_stats.copy(),
                'validation_rules': {
                    'required_sheets': self.required_sheets,
                    'valid_match_results': list(self.valid_match_results),
                    'valid_brands': list(self.valid_brands),
                    'part_number_patterns': len(self.part_number_patterns),
                    'validation_categories': [
                        'part_type', 'brand', 'part_number', 'sales_data',
                        'output_format', 'batch_consistency'
                    ]
                },
                'quality_metrics': {
                    'total_validations': self.validation_stats['total_rows_validated'],
                    'error_rate': self.validation_stats['validation_errors'] / max(1, self.validation_stats['total_rows_validated']),
                    'warning_rate': self.validation_stats['validation_warnings'] / max(1, self.validation_stats['total_rows_validated'])
                },
                'recommendations': self._generate_validation_recommendations()
            }

            self.validation_stats['last_validation'] = datetime.now().isoformat()

            return report

        except Exception as e:
            return {
                'timestamp': datetime.now().isoformat(),
                'error': f"Error generating validation report: {str(e)}",
                'validation_statistics': self.validation_stats.copy()
            }

    def _generate_validation_recommendations(self) -> List[str]:
        """Generate validation improvement recommendations."""
        recommendations = []

        if self.validation_stats['total_rows_validated'] > 0:
            error_rate = self.validation_stats['validation_errors'] / self.validation_stats['total_rows_validated']
            warning_rate = self.validation_stats['validation_warnings'] / self.validation_stats['total_rows_validated']

            if error_rate > 0.1:  # More than 10% errors
                recommendations.append("High error rate detected - review data input processes")
                recommendations.append("Consider implementing stricter input validation")

            if warning_rate > 0.3:  # More than 30% warnings
                recommendations.append("High warning rate - review data quality standards")
                recommendations.append("Consider automated data cleaning processes")

            if error_rate < 0.01 and warning_rate < 0.05:
                recommendations.append("Data quality is excellent - maintain current standards")

        if not recommendations:
            recommendations.append("Insufficient validation data - run more validations for recommendations")

        return recommendations

if __name__ == "__main__":
    # Test the data validator
    print("Testing Data Validator...")

    validator = DataValidator()

    # Test Excel structure validation
    print("1. Testing Excel structure validation...")
    structure_result = validator.validate_excel_structure()
    print(f"   Structure valid: {structure_result['valid']}")
    print(f"   Errors: {len(structure_result['errors'])}")
    print(f"   Warnings: {len(structure_result['warnings'])}")

    # Test row data validation
    print("2. Testing row data validation...")
    test_row = {
        'part_type': 'ENGINE MOUNT',
        'current_supplier': 'ANCHOR',
        'part_number': '3217',
        'skp_part_number': 'SKM3217',
        'call12': 50
    }

    row_result = validator.validate_row_data(test_row)
    print(f"   Row valid: {row_result['valid']}")
    print(f"   Quality score: {row_result['quality_score']:.3f}")
    print(f"   Errors: {len(row_result['errors'])}")
    print(f"   Warnings: {len(row_result['warnings'])}")

    # Test output validation
    print("3. Testing output validation...")
    test_output = {
        'match_result': 'YES',
        'confidence': 95,
        'match_reason': 'Shared OEM reference',
        'fitment_match': 'YES',
        'desc_match': 'YES'
    }

    output_result = validator.validate_output_data(test_output)
    print(f"   Output valid: {output_result['valid']}")
    print(f"   Errors: {len(output_result['errors'])}")

    # Test validation report
    print("4. Testing validation report...")
    report = validator.get_validation_report()
    print(f"   Report sections: {len(report)}")
    print(f"   Recommendations: {len(report.get('recommendations', []))}")

    print("Data Validator test completed.")