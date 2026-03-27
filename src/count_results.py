"""
Quick stats counter — reads Excel and prints current match counts for all sheets.
No scraping required.

Usage:
    uv run python src/count_results.py
    uv run python src/count_results.py --sheet Anchor
"""

import sys
import argparse
from pathlib import Path
import openpyxl

EXCEL_FILE = Path(__file__).parent.parent / "FISHER SKP INTERCHANGE 20260302.xlsx"

ALL_SHEETS = ["GMB", "Four Seasons ", "SMP", "Anchor", "Dorman"]

VERDICTS = ["YES", "LIKELY", "UNCERTAIN", "NO"]


def count_sheet(ws) -> dict:
    counts = {v: 0 for v in VERDICTS}
    counts["BLANK"] = 0
    counts["total_rows"] = 0
    counts["processed"] = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        counts["total_rows"] += 1
        val = row[9] if len(row) > 9 else None  # col J = index 9
        if val in VERDICTS:
            counts[val] += 1
            counts["processed"] += 1
        else:
            counts["BLANK"] += 1

    return counts


def print_report(sheet_name: str, counts: dict):
    processed = counts["processed"]
    total = counts["total_rows"]
    yes = counts["YES"]
    likely = counts["LIKELY"]
    uncertain = counts["UNCERTAIN"]
    no = counts["NO"]

    confirmed = yes + likely
    pct = lambda n: f"{n / processed * 100:.1f}%" if processed > 0 else "N/A"

    print(f"\n{'=' * 55}")
    print(f"  Sheet: {sheet_name}")
    print(f"{'=' * 55}")
    print(f"  Total rows:      {total:>6,}")
    print(f"  Processed rows:  {processed:>6,}   ({pct(processed)} of total)")
    print(f"  Unprocessed:     {counts['BLANK']:>6,}")
    print(f"")
    print(f"  Match Results:")
    print(f"    YES            {yes:>6,}   ({pct(yes)})")
    print(f"    LIKELY         {likely:>6,}   ({pct(likely)})")
    print(f"    UNCERTAIN      {uncertain:>6,}   ({pct(uncertain)})")
    print(f"    NO             {no:>6,}   ({pct(no)})")
    print(f"")
    print(f"  Confirmed match: {confirmed:>6,}   ({pct(confirmed)})  [YES + LIKELY]")
    print(f"  Needs review:    {uncertain:>6,}   ({pct(uncertain)})  [UNCERTAIN]")
    print(f"  Non-match:       {no:>6,}   ({pct(no)})")


def main():
    parser = argparse.ArgumentParser(description="Count match results in the parts Excel file")
    parser.add_argument("--sheet", help="Single sheet to count (e.g. Anchor, Dorman)")
    args = parser.parse_args()

    if not EXCEL_FILE.exists():
        print(f"ERROR: Excel file not found at {EXCEL_FILE}")
        sys.exit(1)

    print(f"Reading: {EXCEL_FILE.name}")
    print(f"Loading workbook (read-only)...")

    wb = openpyxl.load_workbook(str(EXCEL_FILE), read_only=True, data_only=True)

    sheets_to_check = [args.sheet] if args.sheet else ALL_SHEETS
    grand_totals = {v: 0 for v in VERDICTS}
    grand_totals["BLANK"] = 0
    grand_totals["processed"] = 0
    grand_totals["total_rows"] = 0

    for sheet_name in sheets_to_check:
        if sheet_name not in wb.sheetnames:
            print(f"\n  WARNING: Sheet '{sheet_name}' not found, skipping.")
            continue
        ws = wb[sheet_name]
        counts = count_sheet(ws)
        print_report(sheet_name, counts)

        if not args.sheet:  # accumulate for grand total
            for k in grand_totals:
                grand_totals[k] += counts.get(k, 0)

    wb.close()

    # Grand total across all sheets
    if not args.sheet and len(sheets_to_check) > 1:
        print(f"\n{'=' * 55}")
        print(f"  GRAND TOTAL — All Sheets")
        print(f"{'=' * 55}")
        p = grand_totals["processed"]
        pct = lambda n: f"{n / p * 100:.1f}%" if p > 0 else "N/A"
        print(f"  Total rows:      {grand_totals['total_rows']:>6,}")
        print(f"  Processed rows:  {p:>6,}")
        confirmed = grand_totals["YES"] + grand_totals["LIKELY"]
        print(f"  YES + LIKELY:    {confirmed:>6,}   ({pct(confirmed)})")
        print(f"  UNCERTAIN:       {grand_totals['UNCERTAIN']:>6,}   ({pct(grand_totals['UNCERTAIN'])})")
        print(f"  NO:              {grand_totals['NO']:>6,}   ({pct(grand_totals['NO'])})")
        print(f"{'=' * 55}\n")


def count_all_results(excel_path: str = None) -> dict:
    """
    Count results from all sheets and return as a dictionary.
    Used by the unified GUI for statistics display.
    """
    excel_file = Path(excel_path) if excel_path else EXCEL_FILE

    if not excel_file.exists():
        return {}

    try:
        wb = openpyxl.load_workbook(str(excel_file), read_only=True, data_only=True)
        all_results = {}

        for sheet_name in ALL_SHEETS:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                counts = count_sheet(ws)
                all_results[sheet_name] = {
                    "YES": counts["YES"],
                    "LIKELY": counts["LIKELY"],
                    "UNCERTAIN": counts["UNCERTAIN"],
                    "NO": counts["NO"],
                    "total_rows": counts["total_rows"],
                    "processed": counts["processed"]
                }

        wb.close()
        return all_results

    except Exception as e:
        print(f"Error counting results: {e}")
        return {}


if __name__ == "__main__":
    main()
