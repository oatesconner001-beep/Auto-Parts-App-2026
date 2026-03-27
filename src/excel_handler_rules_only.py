"""
Rules-only version of excel_handler.py — no Chrome scraping.

Bypasses Chrome scraping entirely, providing minimal data structures to the
rule-based comparison engine. This avoids Chrome session issues and Excel
corruption while still processing parts based on part numbers.
"""

import os
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rule_compare import compare_parts

# Column indices (0-based)
COL_PART_TYPE   = 0   # A
COL_SUPPLIER    = 1   # B
COL_PART_NUM    = 2   # C
COL_SKP_NUM     = 5   # F
COL_NOTES       = 8   # I

# Output columns (0-based, J=9 through P=15)
COL_MATCH_RESULT  = 9
COL_CONFIDENCE    = 10
COL_MATCH_REASON  = 11
COL_FITMENT_MATCH = 12
COL_DESC_MATCH    = 13
COL_MISSING_INFO  = 14
COL_LAST_CHECKED  = 15

OUTPUT_HEADERS = [
    "MATCH RESULT",     # J
    "CONFIDENCE %",     # K
    "MATCH REASON",     # L
    "FITMENT MATCH",    # M
    "DESC MATCH",       # N
    "MISSING INFO",     # O
    "LAST CHECKED",     # P
]

# Fill colors for MATCH RESULT column
FILL_YES       = PatternFill("solid", fgColor="00FF00")
FILL_LIKELY    = PatternFill("solid", fgColor="90EE90")
FILL_UNCERTAIN = PatternFill("solid", fgColor="FFFF00")
FILL_NO        = PatternFill("solid", fgColor="FF0000")

RESULT_FILLS = {
    "YES":       FILL_YES,
    "LIKELY":    FILL_LIKELY,
    "UNCERTAIN": FILL_UNCERTAIN,
    "NO":        FILL_NO,
}

_BLANK_VALUES = {"", "n/a", "na", "none", "-", "--", "0", ".", "tbd", "?"}


def _is_blank(cell_value) -> bool:
    if cell_value is None:
        return True
    return str(cell_value).strip().lower() in _BLANK_VALUES


def _ensure_headers(sheet):
    """Write output column headers in row 1 if not already present."""
    for i, header in enumerate(OUTPUT_HEADERS):
        col = COL_MATCH_RESULT + i + 1  # openpyxl is 1-based
        cell = sheet.cell(row=1, column=col)
        if not cell.value:
            cell.value = header


def _create_minimal_data(part_number: str, brand: str) -> dict:
    """Create minimal data structure for rules-only comparison."""
    return {
        "found": False,
        "category": "",
        "oem_refs": [],
        "price": "",
        "moreinfo_url": "",
        "image_url": "",
        "specs": "",
        "description": "",
        "features": "",
        "warranty": "",
        "error": "Rules-only mode - no scraping",
        "part_number": part_number,
        "brand": brand
    }


def _write_result(sheet, row_num: int, comparison: dict):
    """Write the 7 output fields to columns J–P for the given row."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    values = [
        comparison.get("match_result", "UNCERTAIN"),
        comparison.get("confidence", 0),
        comparison.get("match_reason", ""),
        comparison.get("fitment_match", "UNKNOWN"),
        comparison.get("desc_match", "UNKNOWN"),
        comparison.get("missing_info", "Rules-only mode"),
        timestamp,
    ]

    for i, value in enumerate(values):
        col = COL_MATCH_RESULT + i + 1
        cell = sheet.cell(row=row_num, column=col)
        cell.value = value

    # Apply color coding to MATCH RESULT cell
    match_result = comparison.get("match_result", "UNCERTAIN")
    if match_result in RESULT_FILLS:
        result_cell = sheet.cell(row=row_num, column=COL_MATCH_RESULT + 1)
        result_cell.fill = RESULT_FILLS[match_result]


def log(msg: str):
    """Print with timestamp."""
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp}] {msg}")


def get_valid_rows(sheet, brand_filter: str, limit: int = None):
    """Extract rows where col B matches brand and both part nums are non-blank."""
    valid_rows = []

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if limit and len(valid_rows) >= limit:
            break

        supplier = str(row[COL_SUPPLIER] or "").strip()
        part_num = str(row[COL_PART_NUM] or "").strip()
        skp_num = str(row[COL_SKP_NUM] or "").strip()
        part_type = str(row[COL_PART_TYPE] or "").strip()

        if (supplier.upper() == brand_filter.upper() and
            not _is_blank(part_num) and not _is_blank(skp_num)):

            valid_rows.append({
                "row_num": row_num,
                "part_num": part_num,
                "skp_num": skp_num,
                "part_type": part_type
            })

    return valid_rows


def process_sheet_rules_only(
    sheet_name: str,
    brand: str,
    filepath: str = "FISHER SKP INTERCHANGE 20260302.xlsx",
    reprocess_uncertain: bool = False,
    limit: int = None,
    on_progress=None,
    stop_flag=None
):
    """
    Process a sheet using only rule-based comparison - no Chrome scraping.

    Args:
        sheet_name: Name of the Excel sheet to process
        brand: Brand filter (only process rows where col B matches this)
        filepath: Path to Excel file
        reprocess_uncertain: Whether to reprocess UNCERTAIN results
        limit: Maximum number of rows to process
        on_progress: Callback function for progress updates
        stop_flag: List with boolean flag for early termination
    """
    log(f"Starting {brand} rules-based processing...")

    # Read valid rows
    wb_read = load_workbook(filepath, read_only=True)
    sheet_read = wb_read[sheet_name]
    valid_rows = get_valid_rows(sheet_read, brand, limit)
    wb_read.close()

    if not valid_rows:
        log("No valid rows found to process.")
        return

    total = len(valid_rows)
    log(f"Found {total} valid rows to process")

    # Open workbook in write mode
    wb = load_workbook(filepath)
    sheet = wb[sheet_name]
    _ensure_headers(sheet)

    try:
        for idx, row_info in enumerate(valid_rows, start=1):
            if stop_flag and stop_flag[0]:
                log("Stop requested — halting after saving.")
                break

            row_num   = row_info["row_num"]
            part_num  = row_info["part_num"]
            skp_num   = row_info["skp_num"]
            part_type = row_info["part_type"]

            # Skip logic: always skip confirmed results; skip UNCERTAIN unless reprocess flag set
            existing = str(sheet.cell(row=row_num, column=COL_MATCH_RESULT + 1).value or "").strip().upper()
            if existing and existing not in ("UNCERTAIN", "") and not reprocess_uncertain:
                log(f"[{idx}/{total}] Row {row_num}: {existing} — skipping")
                if on_progress:
                    on_progress(idx, total)
                continue
            if existing and existing not in ("UNCERTAIN", "") and reprocess_uncertain:
                log(f"[{idx}/{total}] Row {row_num}: already {existing} — skipping (confirmed result)")
                if on_progress:
                    on_progress(idx, total)
                continue
            if existing == "UNCERTAIN" and not reprocess_uncertain:
                log(f"[{idx}/{total}] Row {row_num}: UNCERTAIN — skipping (reprocess not enabled)")
                if on_progress:
                    on_progress(idx, total)
                continue

            log(f"[{idx}/{total}] Row {row_num}: {part_type} | {brand}={part_num} | SKP={skp_num}")

            try:
                # Create minimal data structures for rules-only comparison
                part_data = _create_minimal_data(part_num, brand)
                skp_data = _create_minimal_data(skp_num, "SKP")

                # Run rule-based comparison only
                comparison = compare_parts(part_data, skp_data, part_type)

            except Exception as e:
                log(f"  ERROR on row {row_num}: {e}")
                comparison = {
                    "match_result":  "UNCERTAIN",
                    "confidence":    0,
                    "match_reason":  f"Processing error: {e}",
                    "fitment_match": "UNKNOWN",
                    "desc_match":    "UNKNOWN",
                    "missing_info":  str(e),
                }

            _write_result(sheet, row_num, comparison)
            # Sanitize Unicode characters for Windows console logging
            match_reason = comparison['match_reason'][:80].encode('ascii', errors='replace').decode('ascii')
            log(f"  -> {comparison['match_result']} ({comparison['confidence']}%) - {match_reason}")

            # Crash-safe save: write to temp then replace
            tmp = filepath + ".tmp"
            wb.save(tmp)
            os.replace(tmp, filepath)

            if on_progress:
                on_progress(idx, total)

            # Minimal delay since no network requests
            if idx < total:
                time.sleep(0.1)

    finally:
        tmp = filepath + ".tmp"
        wb.save(tmp)
        os.replace(tmp, filepath)
        wb.close()
        log("Done. Workbook saved.")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Rules-only Excel processing")
    parser.add_argument("sheet", help="Sheet name (Anchor, Dorman, GMB, SMP)")
    parser.add_argument("--limit", type=int, help="Max rows to process")
    parser.add_argument("--reprocess", action="store_true", help="Reprocess UNCERTAIN results")

    args = parser.parse_args()

    # Map sheet names to brands
    sheet_brands = {
        "Anchor": "ANCHOR",
        "Dorman": "DORMAN",
        "GMB": "GMB",
        "SMP": "SMP"
    }

    if args.sheet not in sheet_brands:
        print(f"Error: Unknown sheet '{args.sheet}'. Use: {list(sheet_brands.keys())}")
        exit(1)

    brand = sheet_brands[args.sheet]

    process_sheet_rules_only(
        sheet_name=args.sheet,
        brand=brand,
        reprocess_uncertain=args.reprocess,
        limit=args.limit
    )