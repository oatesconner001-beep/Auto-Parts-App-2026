"""
PHASE 1: Bulk Image URL Scraper — Extract all image URLs first without comparison.

Optimized approach:
- Single browser session for entire batch
- Cache all image URLs to JSON file
- No visual comparison (that's Phase 2)
- Drastically reduce browser session conflicts
"""

import sys
import json
import time
from pathlib import Path
from datetime import datetime
import openpyxl

# Import existing scraper
_src = Path(__file__).parent
if str(_src) not in sys.path:
    sys.path.insert(0, str(_src))

from scraper_headless import scrape_rockauto_subprocess  # HEADLESS VERSION

EXCEL_FILE = Path(__file__).parent.parent / "FISHER SKP INTERCHANGE 20260302.xlsx"
CACHE_DIR = Path(__file__).parent.parent / "image_cache"
CACHE_DIR.mkdir(exist_ok=True)

def _skip_value(value):
    """Check if part number should be skipped"""
    if value is None:
        return True
    s = str(value).strip().upper()
    return s in ["", "N/A", "-", "0", "NONE", "TBD", "?"]

def collect_uncertain_rows(sheet_name):
    """Collect all UNCERTAIN rows that need image analysis"""
    print(f"Loading {sheet_name} UNCERTAIN rows...")

    wb = openpyxl.load_workbook(str(EXCEL_FILE), read_only=True)
    ws = wb[sheet_name]

    brand_filter = sheet_name.upper()
    if brand_filter == "FOUR SEASONS":
        brand_filter = "FOUR SEASONS"

    uncertain_rows = []
    for row_num in range(2, ws.max_row + 1):
        supplier = ws.cell(row=row_num, column=2).value
        match_result = ws.cell(row=row_num, column=10).value  # col J
        part_num = ws.cell(row=row_num, column=3).value       # col C
        skp_num = ws.cell(row=row_num, column=6).value        # col F
        part_type = ws.cell(row=row_num, column=1).value      # col A

        if str(supplier or "").strip().upper() != brand_filter:
            continue
        if match_result != "UNCERTAIN":
            continue
        if _skip_value(part_num) or _skip_value(skp_num):
            continue

        uncertain_rows.append({
            "row_num": row_num,
            "part_num": str(part_num).strip(),
            "skp_num": str(skp_num).strip(),
            "part_type": str(part_type or "").strip(),
        })

    wb.close()
    print(f"Found {len(uncertain_rows):,} UNCERTAIN rows")
    return uncertain_rows

def bulk_scrape_images(sheet_name, limit=None):
    """Phase 1: Scrape all image URLs and cache them"""
    cache_file = CACHE_DIR / f"{sheet_name.lower()}_image_urls.json"

    # Load existing cache if available
    cached_data = {}
    if cache_file.exists():
        print(f"Loading existing cache: {cache_file}")
        try:
            with open(cache_file, 'r') as f:
                cached_data = json.load(f)
            print(f"Found {len(cached_data)} cached entries")
        except:
            print("Cache file corrupted, starting fresh")
            cached_data = {}

    # Collect rows to process
    uncertain_rows = collect_uncertain_rows(sheet_name)
    if limit:
        uncertain_rows = uncertain_rows[:limit]

    print(f"\n🔍 BULK SCRAPING: {len(uncertain_rows)} rows")
    print(f"Cache file: {cache_file}")
    print(f"[Vision] Using subprocess scraper (isolated Chrome sessions)")

    processed = 0
    found_images = 0
    errors = 0

    start_time = time.time()

    for i, row in enumerate(uncertain_rows, 1):
        part_key = f"{row['part_num']}|{row['skp_num']}"

        # Skip if already cached
        if part_key in cached_data:
            if cached_data[part_key].get('found_images', 0) > 0:
                found_images += 1
            processed += 1
            continue

        print(f"[{i}/{len(uncertain_rows)}] Scraping: {row['part_num']} vs {row['skp_num']}")

        # Scrape both parts
        brand_name = sheet_name
        result_data = {
            'part_num': row['part_num'],
            'skp_num': row['skp_num'],
            'part_type': row['part_type'],
            'row_num': row['row_num'],
            'scraped_at': datetime.now().isoformat(),
            'brand_image_url': None,
            'skp_image_url': None,
            'found_images': 0,
            'error': None
        }

        try:
            # Scrape brand part
            print(f"  Scraping {brand_name} {row['part_num']}...")
            brand_result = scrape_rockauto_subprocess(row['part_num'], brand_name)
            if brand_result.get('found') and brand_result.get('image_url'):
                result_data['brand_image_url'] = brand_result['image_url']
                result_data['found_images'] += 1
                print(f"    [OK] Found {brand_name} image")
            else:
                print(f"    [FAIL] No {brand_name} image")

            # Small delay to avoid overwhelming RockAuto
            time.sleep(0.5)

            # Scrape SKP part
            print(f"  Scraping SKP {row['skp_num']}...")
            skp_result = scrape_rockauto_subprocess(row['skp_num'], 'SKP')
            if skp_result.get('found') and skp_result.get('image_url'):
                result_data['skp_image_url'] = skp_result['image_url']
                result_data['found_images'] += 1
                print(f"    [OK] Found SKP image")
            else:
                print(f"    [FAIL] No SKP image")

        except Exception as e:
            result_data['error'] = str(e)
            print(f"    ⚠️ Scrape error: {e}")
            errors += 1

        # Cache result
        cached_data[part_key] = result_data
        processed += 1

        if result_data['found_images'] > 0:
            found_images += 1

        # Save cache every 10 rows
        if processed % 10 == 0:
            with open(cache_file, 'w') as f:
                json.dump(cached_data, f, indent=2)

            elapsed = time.time() - start_time
            rate = processed / elapsed * 60  # per minute
            print(f"    -- Progress: {processed}/{len(uncertain_rows)} | Images: {found_images} | Rate: {rate:.1f}/min")

    # Final save
    with open(cache_file, 'w') as f:
        json.dump(cached_data, f, indent=2)

    elapsed = time.time() - start_time
    print(f"\n[OK] BULK SCRAPING COMPLETE")
    print(f"   Processed: {processed:,} rows")
    print(f"   Found images: {found_images:,} ({found_images/processed*100:.1f}%)")
    print(f"   Errors: {errors:,}")
    print(f"   Time: {elapsed:.1f}s ({processed/elapsed*60:.1f} rows/min)")
    print(f"   Cache: {cache_file}")

    return cache_file, processed, found_images

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("sheet", choices=["Anchor", "Dorman", "GMB", "SMP", "Four Seasons"], help="Sheet to process")
    parser.add_argument("--limit", type=int, help="Limit number of rows")

    args = parser.parse_args()

    bulk_scrape_images(args.sheet, args.limit)