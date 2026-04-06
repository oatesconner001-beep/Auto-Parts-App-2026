"""
Excel reader/writer for the Parts Agent.

Reads valid rows from a named sheet, calls the scraper + rule/AI compare,
and writes 7 result columns (J–P) back to the file with color coding.

Each sheet is processed for its own brand:
  - 'Anchor' sheet  → search RockAuto for ANCHOR brand
  - 'Dorman' sheet  → search RockAuto for DORMAN brand
  - etc.

Only rows where col B (CURRENT SUPPLIER) matches the target brand are processed.
"""

import os
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Column indices (0-based)
COL_PART_TYPE   = 0   # A
COL_SUPPLIER    = 1   # B
COL_PART_NUM    = 2   # C
COL_SKP_NUM     = 5   # F
COL_NOTES       = 8   # I

# Output columns (0-based, J=9 through P=15)
COL_MATCH_RESULT  = 9
COL_CONFIDENCE    = 10
COL_MATCH_REASON  = 11
COL_FITMENT_MATCH = 12
COL_DESC_MATCH    = 13
COL_MISSING_INFO  = 14
COL_LAST_CHECKED  = 15

OUTPUT_HEADERS = [
    "MATCH RESULT",     # J
    "CONFIDENCE %",     # K
    "MATCH REASON",     # L
    "FITMENT MATCH",    # M
    "DESC MATCH",       # N
    "MISSING INFO",     # O
    "LAST CHECKED",     # P
]

# Fill colors for MATCH RESULT column
FILL_YES       = PatternFill("solid", fgColor="00FF00")
FILL_LIKELY    = PatternFill("solid", fgColor="90EE90")
FILL_UNCERTAIN = PatternFill("solid", fgColor="FFFF00")
FILL_NO        = PatternFill("solid", fgColor="FF0000")

RESULT_FILLS = {
    "YES":       FILL_YES,
    "LIKELY":    FILL_LIKELY,
    "UNCERTAIN": FILL_UNCERTAIN,
    "NO":        FILL_NO,
}

# Delay between rows (seconds) to avoid Firecrawl rate limiting
ROW_DELAY = 2.5

# Sites to query for multi-site enrichment (excludes RockAuto — scraped separately)
_ENRICH_SITES = ['ACDelco', 'PartsGeek', 'ShowMeTheParts']

_BLANK_VALUES = {"", "n/a", "na", "none", "-", "--", "0", ".", "tbd", "?"}


def _is_blank(cell_value) -> bool:
    if cell_value is None:
        return True
    return str(cell_value).strip().lower() in _BLANK_VALUES


def _ensure_headers(sheet):
    """Write output column headers in row 1 if not already present."""
    for i, header in enumerate(OUTPUT_HEADERS):
        col = COL_MATCH_RESULT + i + 1  # openpyxl is 1-based
        cell = sheet.cell(row=1, column=col)
        if not cell.value:
            cell.value = header


def _already_processed(sheet, row_num: int) -> bool:
    """Return True if the MATCH RESULT cell already has a non-UNCERTAIN value."""
    cell = sheet.cell(row=row_num, column=COL_MATCH_RESULT + 1)
    val = str(cell.value or "").strip().upper()
    return bool(val) and val != "UNCERTAIN"


def _write_result(sheet, row_num: int, comparison: dict):
    """Write the 7 output fields to columns J–P for the given row."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    values = [
        comparison.get("match_result", "UNCERTAIN"),
        comparison.get("confidence", 0),
        comparison.get("match_reason", ""),
        comparison.get("fitment_match", "UNKNOWN"),
        comparison.get("desc_match", "UNKNOWN"),
        comparison.get("missing_info", ""),
        timestamp,
    ]

    for i, val in enumerate(values):
        col = COL_MATCH_RESULT + i + 1  # 1-based
        sheet.cell(row=row_num, column=col).value = val

    # Color code the MATCH RESULT cell (col J)
    match_result = comparison.get("match_result", "UNCERTAIN")
    fill = RESULT_FILLS.get(match_result, FILL_UNCERTAIN)
    sheet.cell(row=row_num, column=COL_MATCH_RESULT + 1).fill = fill


def get_valid_rows(filepath: str, sheet_name: str, supplier_filter: str = None) -> list[dict]:
    """
    Read a sheet and return rows where both col C and col F are non-blank.
    If supplier_filter is given, only return rows where col B matches it (case-insensitive).

    Each dict: row_num, part_type, supplier, part_num, skp_num, notes
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheet = wb[sheet_name]
    rows = []
    supplier_upper = supplier_filter.upper().strip() if supplier_filter else None

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Handle sheets with fewer than 10 columns
        part_num = row[COL_PART_NUM] if len(row) > COL_PART_NUM else None
        skp_num  = row[COL_SKP_NUM]  if len(row) > COL_SKP_NUM  else None
        supplier = str(row[COL_SUPPLIER] or "").strip() if len(row) > COL_SUPPLIER else ""

        if _is_blank(part_num) or _is_blank(skp_num):
            continue

        if supplier_upper and supplier.upper() != supplier_upper:
            continue

        rows.append({
            "row_num":   row_num,
            "part_type": str(row[COL_PART_TYPE] or "").strip() if len(row) > COL_PART_TYPE else "",
            "supplier":  supplier,
            "part_num":  str(part_num).strip(),
            "skp_num":   str(skp_num).strip(),
            "notes":     str(row[COL_NOTES] or "").strip() if len(row) > COL_NOTES and row[COL_NOTES] else "",
        })

    wb.close()
    return rows


def _normalize_oem_ref(ref) -> str:
    """Extract plain OEM number string from either str or dict format."""
    if isinstance(ref, dict):
        return ref.get('oem_number', '')
    return str(ref)


def _fitment_to_features(fitment_data: list) -> list:
    """Convert structured fitment_data to text strings for rule_compare's regex matching."""
    texts = set()
    for f in (fitment_data or []):
        make = f.get('make', '')
        model = f.get('model', '')
        yr = f.get('year') or f.get('year_start')
        yr_end = f.get('year_end')
        if make and yr:
            if yr_end and yr_end != yr:
                texts.add(f"{yr}-{yr_end} {make} {model}".strip())
            else:
                texts.add(f"{yr} {make} {model}".strip())
    return list(texts)


def _enrich_with_multi_site(part_data: dict, part_number: str, brand: str,
                            mgr, log) -> dict:
    """Merge OEM refs, category, fitment, specs, description from non-RockAuto sites.

    Returns a new dict (same shape as part_data) with enriched fields.
    Falls back to unmodified part_data on any error.
    """
    try:
        ms_result = mgr.scrape_part_multi_site(
            part_number, brand,
            sites=_ENRICH_SITES,
            store_results=True,
        )
    except Exception as e:
        log(f"  Multi-site scrape failed: {e}")
        return part_data

    found_sites = ms_result.get('summary', {}).get('found_on_sites', 0)
    if found_sites == 0:
        return part_data

    enriched = dict(part_data)  # shallow copy

    # --- Merge OEM refs (deduplicated, normalized to plain strings) ---
    existing_oems = {_normalize_oem_ref(r).upper() for r in (enriched.get('oem_refs') or []) if r}
    merged_oems = [_normalize_oem_ref(r) for r in (enriched.get('oem_refs') or []) if r]

    for site_name, site_data in ms_result.get('sites', {}).items():
        if not site_data.get('found'):
            continue
        for ref in (site_data.get('oem_refs') or []):
            norm = _normalize_oem_ref(ref)
            if norm and norm.upper() not in existing_oems:
                merged_oems.append(norm)
                existing_oems.add(norm.upper())

    enriched['oem_refs'] = merged_oems

    # --- Best category (fill if RockAuto's is None/empty) ---
    if not enriched.get('category'):
        for site_name, site_data in ms_result.get('sites', {}).items():
            if site_data.get('found') and site_data.get('category'):
                enriched['category'] = site_data['category']
                break

    # --- Best description (fill if RockAuto's is None/empty) ---
    if not enriched.get('description'):
        for site_name, site_data in ms_result.get('sites', {}).items():
            if site_data.get('found') and site_data.get('description'):
                enriched['description'] = site_data['description']
                break

    # --- Merge specs (RockAuto keys take precedence) ---
    for site_name, site_data in ms_result.get('sites', {}).items():
        if not site_data.get('found'):
            continue
        for k, v in (site_data.get('specs') or {}).items():
            if k not in (enriched.get('specs') or {}):
                if enriched.get('specs') is None:
                    enriched['specs'] = {}
                enriched['specs'][k] = v

    # --- Merge fitment into features (for rule_compare's regex-based fitment scoring) ---
    extra_features = []
    for site_name, site_data in ms_result.get('sites', {}).items():
        if not site_data.get('found'):
            continue
        extra_features.extend(_fitment_to_features(site_data.get('fitment_data')))

    if extra_features:
        existing_features = enriched.get('features') or []
        enriched['features'] = existing_features + extra_features

    site_names = [s for s, d in ms_result.get('sites', {}).items() if d.get('found')]
    log(f"  Enriched from {len(site_names)} site(s): {', '.join(site_names)} "
        f"(+{len(merged_oems) - len(part_data.get('oem_refs') or [])} OEM refs)")

    return enriched


def process_rows(
    filepath: str,
    sheet_name: str = "Anchor",
    search_brand: str = None,
    reprocess_uncertain: bool = False,
    on_progress=None,
    on_log=None,
    stop_flag=None,
    start_row: int = None,
    end_row: int = None,
    limit: int = None,
):
    """
    Main processing loop.

    Args:
        filepath:             Path to the Excel file
        sheet_name:           Which sheet to read and write (e.g. 'Anchor', 'Dorman')
        search_brand:         Brand to search on RockAuto for the left-side part.
                              Defaults to sheet_name if not provided.
        reprocess_uncertain:  If True, rows already marked UNCERTAIN will be re-processed.
        on_progress:          Callback(current, total)
        on_log:               Callback(message: str)
        stop_flag:            [bool] — set [0]=True to stop after current row
        start_row:            Optional first spreadsheet row number to process
        end_row:              Optional last spreadsheet row number to process
    """
    from scraper_subprocess import scrape_rockauto_subprocess as scrape_rockauto
    from rule_compare import compare_parts

    brand = (search_brand or sheet_name).upper().strip()

    def log(msg):
        if on_log:
            on_log(msg)
        else:
            print(msg)

    log(f"Sheet: {sheet_name} | Brand: {brand} | Reprocess UNCERTAIN: {reprocess_uncertain}")
    log(f"Loading workbook: {filepath}")

    # Initialize multi-site enrichment (graceful fallback to RockAuto-only)
    multi_site_mgr = None
    try:
        from scrapers.multi_site_manager import MultiSiteScraperManager
        multi_site_mgr = MultiSiteScraperManager()
        log(f"Multi-site enrichment enabled ({len(_ENRICH_SITES)} sites)")
    except Exception as e:
        log(f"Multi-site enrichment unavailable (RockAuto-only mode): {e}")

    # Only load rows whose col B matches the brand
    valid_rows = get_valid_rows(filepath, sheet_name, supplier_filter=brand)
    log(f"Found {len(valid_rows)} rows for {brand} brand")

    # Filter by row range
    if start_row is not None:
        valid_rows = [r for r in valid_rows if r["row_num"] >= start_row]
    if end_row is not None:
        valid_rows = [r for r in valid_rows if r["row_num"] <= end_row]
    if limit is not None:
        valid_rows = valid_rows[:limit]

    total = len(valid_rows)

    # Open workbook in write mode
    wb = load_workbook(filepath)
    sheet = wb[sheet_name]
    _ensure_headers(sheet)

    consec_errors = 0  # consecutive scraper/compare failure counter

    try:
        for idx, row_info in enumerate(valid_rows, start=1):
            if stop_flag and stop_flag[0]:
                log("Stop requested — halting after saving.")
                break

            row_num   = row_info["row_num"]
            part_num  = row_info["part_num"]
            skp_num   = row_info["skp_num"]
            part_type = row_info["part_type"]

            # Skip logic: always skip confirmed results; skip UNCERTAIN unless reprocess flag set
            existing = str(sheet.cell(row=row_num, column=COL_MATCH_RESULT + 1).value or "").strip().upper()
            if existing and existing not in ("UNCERTAIN", "") and not reprocess_uncertain:
                log(f"[{idx}/{total}] Row {row_num}: {existing} — skipping")
                if on_progress:
                    on_progress(idx, total)
                continue
            if existing and existing not in ("UNCERTAIN", "") and reprocess_uncertain:
                log(f"[{idx}/{total}] Row {row_num}: already {existing} — skipping (confirmed result)")
                if on_progress:
                    on_progress(idx, total)
                continue
            if existing == "UNCERTAIN" and not reprocess_uncertain:
                log(f"[{idx}/{total}] Row {row_num}: UNCERTAIN — skipping (reprocess not enabled)")
                if on_progress:
                    on_progress(idx, total)
                continue

            log(f"[{idx}/{total}] Row {row_num}: {part_type} | {brand}={part_num} | SKP={skp_num}")

            comparison = None
            _last_error = None
            for _attempt in range(3):
                try:
                    part_data = scrape_rockauto(part_num, brand=brand)
                    skp_data  = scrape_rockauto(skp_num,  brand="SKP")

                    # Enrich with cross-site data if manager is available
                    if multi_site_mgr is not None:
                        part_data = _enrich_with_multi_site(part_data, part_num, brand, multi_site_mgr, log)
                        skp_data  = _enrich_with_multi_site(skp_data, skp_num, "SKP", multi_site_mgr, log)

                    comparison = compare_parts(part_data, skp_data, part_type)
                    break
                except Exception as e:
                    _last_error = e
                    if _attempt < 2:
                        _delay = 2 * (2 ** _attempt)  # 2s, 4s
                        log(f"  Retry {_attempt + 1}/3 on row {row_num} (after {_delay}s): {e}")
                        time.sleep(_delay)

            if comparison is None:
                log(f"  ERROR on row {row_num} after 3 attempts: {_last_error}")
                consec_errors += 1
                if consec_errors >= 5:
                    log(f"  WARNING: 5 consecutive errors — pausing 30s to allow recovery")
                    time.sleep(30)
                    consec_errors = 0
                comparison = {
                    "match_result":  "UNCERTAIN",
                    "confidence":    0,
                    "match_reason":  f"Processing error: {_last_error}",
                    "fitment_match": "UNKNOWN",
                    "desc_match":    "UNKNOWN",
                    "missing_info":  str(_last_error),
                }
            else:
                consec_errors = 0

            _write_result(sheet, row_num, comparison)
            # Sanitize Unicode characters for Windows console logging
            match_reason = comparison['match_reason'][:80].encode('ascii', errors='replace').decode('ascii')
            log(f"  -> {comparison['match_result']} ({comparison['confidence']}%) - {match_reason}")

            # Crash-safe save: write to temp then replace
            tmp = filepath + ".tmp"
            wb.save(tmp)
            os.replace(tmp, filepath)

            if on_progress:
                on_progress(idx, total)

            if idx < total:
                time.sleep(ROW_DELAY)

    finally:
        tmp = filepath + ".tmp"
        wb.save(tmp)
        os.replace(tmp, filepath)
        wb.close()
        log("Done. Workbook saved.")


if __name__ == "__main__":
    filepath = "FISHER SKP INTERCHANGE 20260302.xlsx"

    import sys
    import argparse
    parser = argparse.ArgumentParser(description="Process Excel sheet for parts matching")
    parser.add_argument("sheet", nargs="?", default="Anchor", help="Sheet name to process")
    parser.add_argument("--reprocess", action="store_true", help="Reprocess UNCERTAIN rows")
    parser.add_argument("--limit", type=int, default=None, help="Process only first N eligible rows")
    args = parser.parse_args()

    sheet = args.sheet
    reprocess = args.reprocess
    limit = args.limit

    print(f"=== Processing sheet: {sheet} | reprocess_uncertain={reprocess} | limit={limit} ===")
    process_rows(
        filepath,
        sheet_name=sheet,
        reprocess_uncertain=reprocess,
        on_log=print,
        limit=limit,
    )
    print("Done.")
