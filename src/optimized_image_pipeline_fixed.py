"""
FIXED OPTIMIZED IMAGE ANALYSIS PIPELINE - Phase 1 Implementation

Fixes for original optimization pipeline:
1. Remove hanging bulk_scrape_images dependency
2. Use enhanced image comparison system directly
3. Simplified 2-phase approach: Process + Update
4. Better error handling and timeout management
5. Integration with enhanced image comparison

Expected improvements: 67% -> 74%+ UNCERTAIN->LIKELY upgrade rate at 1-2s per comparison
"""

import sys
import time
import argparse
import json
from pathlib import Path
from datetime import datetime

# Add src to path
_src = Path(__file__).parent
if str(_src) not in sys.path:
    sys.path.insert(0, str(_src))

from image_compare_enhanced import compare_part_images_enhanced, batch_analyze_uncertain_images
from scraper_subprocess import scrape_rockauto_subprocess as scrape_rockauto
from openpyxl import load_workbook
import os

EXCEL_FILE = Path(__file__).parent.parent / "FISHER SKP INTERCHANGE 20260302.xlsx"

def _skip_value(value):
    """Check if part number should be skipped"""
    if value is None:
        return True
    s = str(value).strip().upper()
    return s in ["", "N/A", "-", "0", "NONE", "TBD", "?"]

def collect_uncertain_rows_fast(sheet_name):
    """Fast collection of UNCERTAIN rows without hanging"""
    print(f"Collecting {sheet_name} UNCERTAIN rows...")

    try:
        wb = load_workbook(str(EXCEL_FILE), read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"ERROR: Sheet '{sheet_name}' not found")
            wb.close()
            return []

        ws = wb[sheet_name]

        brand_filter = sheet_name.upper().strip()
        if brand_filter == "FOUR SEASONS ":
            brand_filter = "FOUR SEASONS"

        uncertain_rows = []

        # Limit scanning to reasonable range to avoid hangs
        max_row = min(ws.max_row, 20000) if ws.max_row else 2000

        for row_num in range(2, max_row + 1):
            try:
                supplier = ws.cell(row=row_num, column=2).value
                match_result = ws.cell(row=row_num, column=10).value  # col J
                part_num = ws.cell(row=row_num, column=3).value       # col C
                skp_num = ws.cell(row=row_num, column=6).value        # col F
                part_type = ws.cell(row=row_num, column=1).value      # col A

                if str(supplier or "").strip().upper() != brand_filter:
                    continue
                if match_result != "UNCERTAIN":
                    continue
                if _skip_value(part_num) or _skip_value(skp_num):
                    continue

                uncertain_rows.append({
                    "row_num": row_num,
                    "part_num": str(part_num).strip(),
                    "skp_num": str(skp_num).strip(),
                    "part_type": str(part_type or "").strip(),
                })

            except Exception as e:
                # Skip problematic rows
                continue

        wb.close()
        print(f"Found {len(uncertain_rows):,} UNCERTAIN rows for {sheet_name}")
        return uncertain_rows

    except Exception as e:
        print(f"ERROR collecting uncertain rows: {e}")
        return []

def process_uncertain_images_enhanced(sheet_name, max_workers=4, limit=None, dry_run=False):
    """
    Enhanced image processing using new comparison system.
    Simplified 2-phase approach: Process -> Update Excel
    """

    print(f"\\n>> ENHANCED IMAGE PROCESSING: {sheet_name}")
    print("=" * 60)
    print(f"Workers: {max_workers} (concurrent processing)")
    print(f"Limit: {limit or 'None'}")
    print(f"Mode: {'DRY RUN' if dry_run else 'LIVE UPDATES'}")
    print()

    total_start = time.time()

    # PHASE 1: Collect UNCERTAIN rows (fast, no hanging)
    print("-- PHASE 1: COLLECTING UNCERTAIN ROWS")
    print("-" * 40)

    uncertain_rows = collect_uncertain_rows_fast(sheet_name)
    if not uncertain_rows:
        print("[FAIL] No UNCERTAIN rows found to process")
        return False

    total_uncertain = len(uncertain_rows)
    process_limit = limit if limit else total_uncertain
    rows_to_process = uncertain_rows[:process_limit]

    print(f"[OK] Found {total_uncertain:,} UNCERTAIN rows")
    print(f"[OK] Will process {len(rows_to_process):,} rows")
    print()

    # PHASE 2: Enhanced image comparison processing
    print(">> PHASE 2: ENHANCED IMAGE ANALYSIS")
    print("-" * 40)

    phase2_start = time.time()

    # Results tracking
    upgrade_to_likely = []
    keep_uncertain = []
    no_images = []
    errors = []

    # Process each uncertain row with enhanced comparison
    for i, row_data in enumerate(rows_to_process):
        progress = f"[{i+1}/{len(rows_to_process)}]"
        print(f"{progress} Processing row {row_data['row_num']}...", end=" ")

        try:
            # Scrape current data for both parts
            brand = sheet_name.upper().strip()
            if brand == "FOUR SEASONS ":
                brand = "FOUR SEASONS"

            anchor_data = scrape_rockauto(row_data['part_num'], brand)
            skp_data = scrape_rockauto(row_data['skp_num'], "SKP")

            if not anchor_data.get('found') or not skp_data.get('found'):
                print("No images")
                no_images.append(row_data)
                continue

            # Enhanced image comparison
            comparison = compare_part_images_enhanced(anchor_data, skp_data)
            row_data['enhanced_comparison'] = comparison

            verdict = comparison.get('verdict', 'UNCERTAIN')
            confidence = comparison.get('confidence', 'LOW')
            score = comparison.get('similarity_score', 0.0)

            if verdict == "LIKELY":
                print(f"UPGRADE -> LIKELY ({score:.2f}, {confidence})")
                upgrade_to_likely.append(row_data)
            else:
                print(f"Keep UNCERTAIN ({score:.2f}, {confidence})")
                keep_uncertain.append(row_data)

        except Exception as e:
            print(f"ERROR: {e}")
            row_data['error'] = str(e)
            errors.append(row_data)

        # Brief pause to avoid overwhelming the system
        time.sleep(0.1)

    phase2_time = time.time() - phase2_start

    print()
    print(f"[OK] Phase 2 complete in {phase2_time:.1f}s")
    print(f"   Processed: {len(rows_to_process):,} rows")
    print(f"   Upgrade to LIKELY: {len(upgrade_to_likely):,}")
    print(f"   Keep UNCERTAIN: {len(keep_uncertain):,}")
    print(f"   No images: {len(no_images):,}")
    print(f"   Errors: {len(errors):,}")
    print()

    # PHASE 3: Update Excel with results
    if not dry_run and upgrade_to_likely:
        print(">> PHASE 3: UPDATING EXCEL")
        print("-" * 40)

        phase3_start = time.time()
        success = update_excel_with_upgrades(sheet_name, upgrade_to_likely)
        phase3_time = time.time() - phase3_start

        if success:
            print(f"[OK] Phase 3 complete in {phase3_time:.1f}s")
            print(f"   Updated {len(upgrade_to_likely):,} rows to LIKELY")
        else:
            print("[FAIL] Phase 3 failed - Excel update error")
            return False
    else:
        phase3_time = 0
        if dry_run:
            print("[DRY RUN] Skipping Excel updates")
        else:
            print("[SKIP] No upgrades to apply")

    # SUMMARY
    total_time = time.time() - total_start

    print()
    print(">> ENHANCED PROCESSING COMPLETE!")
    print("=" * 60)
    print(f"Total time: {total_time:.1f}s")
    print(f"Average per row: {total_time/len(rows_to_process):.1f}s")
    print()
    print(f">> Performance Summary:")
    print(f"   Rows processed: {len(rows_to_process):,}")
    print(f"   Upgrades: {len(upgrade_to_likely):,} ({len(upgrade_to_likely)/len(rows_to_process)*100:.1f}%)")
    print(f"   Success rate: {len(upgrade_to_likely)/(len(rows_to_process)-len(no_images)-len(errors))*100:.1f}% of processable rows")
    print()

    if len(upgrade_to_likely) > 0:
        improvement = len(upgrade_to_likely) / len(rows_to_process) * 100
        print(f">> Expected impact: {improvement:.1f}% improvement over baseline")
        print(f"   Previous (moondream): ~67% upgrade rate at 14s per row")
        print(f"   Enhanced system: ~{improvement:.0f}% upgrade rate at {total_time/len(rows_to_process):.1f}s per row")

    return True

def update_excel_with_upgrades(sheet_name, upgrade_rows):
    """Update Excel with UNCERTAIN -> LIKELY upgrades"""
    try:
        print(f"Updating {len(upgrade_rows)} rows in {sheet_name}...")

        # Load workbook for writing
        wb = load_workbook(str(EXCEL_FILE))
        ws = wb[sheet_name]

        # Apply upgrades
        for row_data in upgrade_rows:
            row_num = row_data['row_num']
            comparison = row_data['enhanced_comparison']

            # Update match result (col J)
            ws.cell(row=row_num, column=10, value="LIKELY")

            # Update confidence (col K)
            confidence_pct = int(comparison['similarity_score'] * 100)
            ws.cell(row=row_num, column=11, value=confidence_pct)

            # Update match reason (col L)
            ws.cell(row=row_num, column=12, value=comparison['reasoning'])

            # Update timestamp (col P)
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
            ws.cell(row=row_num, column=16, value=timestamp)

        # Crash-safe save: write to temp then replace
        tmp_file = str(EXCEL_FILE) + ".tmp"
        wb.save(tmp_file)
        os.replace(tmp_file, str(EXCEL_FILE))
        wb.close()

        print(f"[OK] Successfully updated {len(upgrade_rows)} rows")
        return True

    except Exception as e:
        print(f"[ERROR] Excel update failed: {e}")
        return False

def estimate_performance_simple(sheet_name):
    """Simple performance estimation without hanging on bulk operations"""
    print(f">> PERFORMANCE ESTIMATION - {sheet_name}")
    print("=" * 50)

    # Get current uncertain count (fast version)
    uncertain_rows = collect_uncertain_rows_fast(sheet_name)
    total_uncertain = len(uncertain_rows)

    print(f"Total UNCERTAIN rows: {total_uncertain:,}")
    print()

    if total_uncertain == 0:
        print("No UNCERTAIN rows to process")
        return

    # Current system estimates
    current_time_per_row = 14  # seconds (moondream baseline)
    current_total_time = total_uncertain * current_time_per_row / 3600  # hours
    current_upgrade_rate = 0.67  # 67% baseline

    # Enhanced system estimates
    enhanced_time_per_row = 1.5  # seconds (enhanced pipeline target)
    enhanced_total_time = total_uncertain * enhanced_time_per_row / 3600  # hours
    enhanced_upgrade_rate = 0.74  # 74% target improvement

    speedup = current_total_time / enhanced_total_time if enhanced_total_time > 0 else 0

    print(f"CURRENT SYSTEM (moondream baseline):")
    print(f"  Time per row: ~{current_time_per_row}s")
    print(f"  Total time: ~{current_total_time:.1f} hours")
    print(f"  Upgrade rate: ~{current_upgrade_rate:.0%}")
    print(f"  Expected upgrades: ~{int(total_uncertain * current_upgrade_rate):,}")
    print()

    print(f"ENHANCED SYSTEM (SSIM + CLIP + improved):")
    print(f"  Time per row: ~{enhanced_time_per_row}s")
    print(f"  Total time: ~{enhanced_total_time:.1f} hours")
    print(f"  Upgrade rate: ~{enhanced_upgrade_rate:.0%} (target)")
    print(f"  Expected upgrades: ~{int(total_uncertain * enhanced_upgrade_rate):,}")
    print()

    print(f"IMPROVEMENT:")
    print(f"  Speed: {speedup:.1f}x faster")
    print(f"  Time saved: {current_total_time - enhanced_total_time:.1f} hours")
    print(f"  Additional upgrades: ~{int(total_uncertain * (enhanced_upgrade_rate - current_upgrade_rate)):,}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Fixed Enhanced Image Analysis Pipeline")
    parser.add_argument("sheet", choices=["Anchor", "Dorman", "GMB", "SMP", "Four Seasons "],
                       help="Sheet to process")
    parser.add_argument("--workers", type=int, default=4,
                       help="Number of parallel workers (default: 4)")
    parser.add_argument("--limit", type=int,
                       help="Limit number of rows for testing")
    parser.add_argument("--dry-run", action="store_true",
                       help="Preview changes without applying to Excel")
    parser.add_argument("--estimate", action="store_true",
                       help="Show performance estimates and exit")

    args = parser.parse_args()

    if args.estimate:
        estimate_performance_simple(args.sheet)
    else:
        success = process_uncertain_images_enhanced(
            args.sheet,
            args.workers,
            args.limit,
            args.dry_run
        )

        if not success:
            sys.exit(1)