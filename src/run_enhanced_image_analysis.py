"""
Enhanced Image Analysis Runner - Phase 1 Implementation

Simpler approach that bypasses the hanging bulk scraper:
1. Use existing count_results.py approach for Excel reading (proven fast)
2. Process UNCERTAIN rows directly with enhanced comparison
3. Apply upgrades immediately to Excel
4. Avoid the complex 3-phase pipeline that hangs

Target: 67% -> 74%+ upgrade rate at 1-2s per comparison
"""

import sys
import time
import argparse
from pathlib import Path
from datetime import datetime

# Add src to path
_src = Path(__file__).parent
if str(_src) not in sys.path:
    sys.path.insert(0, str(_src))

from image_compare_enhanced import compare_part_images_enhanced
from scraper_subprocess import scrape_rockauto_subprocess as scrape_rockauto
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

EXCEL_FILE = Path(__file__).parent.parent / "FISHER SKP INTERCHANGE 20260302.xlsx"

def _skip_value(value):
    """Check if part number should be skipped"""
    if value is None:
        return True
    s = str(value).strip().upper()
    return s in ["", "N/A", "-", "0", "NONE", "TBD", "?"]

def find_uncertain_rows(sheet_name, limit=None):
    """Find UNCERTAIN rows efficiently using proven approach from count_results.py"""
    print(f"Loading {sheet_name} UNCERTAIN rows...")

    wb = load_workbook(str(EXCEL_FILE), read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet '{sheet_name}' not found")
        wb.close()
        return []

    ws = wb[sheet_name]

    brand_filter = sheet_name.upper().strip()
    if brand_filter == "FOUR SEASONS ":
        brand_filter = "FOUR SEASONS"

    uncertain_rows = []

    # Use efficient row iteration like count_results.py
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            # Map to our columns
            part_type = row[0] if len(row) > 0 else None      # A
            supplier = row[1] if len(row) > 1 else None       # B
            part_num = row[2] if len(row) > 2 else None       # C
            skp_num = row[5] if len(row) > 5 else None        # F
            match_result = row[9] if len(row) > 9 else None   # J

            # Filter logic
            if str(supplier or "").strip().upper() != brand_filter:
                continue
            if match_result != "UNCERTAIN":
                continue
            if _skip_value(part_num) or _skip_value(skp_num):
                continue

            # Calculate actual row number (iter_rows is 1-indexed, we start at row 2)
            row_num = len(uncertain_rows) + 2

            uncertain_rows.append({
                "row_num": row_num,
                "part_num": str(part_num).strip(),
                "skp_num": str(skp_num).strip(),
                "part_type": str(part_type or "").strip()
            })

            # Apply limit if specified
            if limit and len(uncertain_rows) >= limit:
                break

        except Exception as e:
            # Skip problematic rows
            continue

    wb.close()
    print(f"Found {len(uncertain_rows)} UNCERTAIN rows in {sheet_name}")
    return uncertain_rows

def run_enhanced_image_analysis(sheet_name, limit=None, dry_run=False):
    """
    Run enhanced image analysis on UNCERTAIN rows.
    Simple, direct approach without complex pipeline.
    """

    print(f"\\n>> ENHANCED IMAGE ANALYSIS: {sheet_name}")
    print("=" * 55)
    print(f"Target: Improve 67% -> 74%+ upgrade rate")
    print(f"Method: SSIM + CLIP + Enhanced comparison")
    print(f"Limit: {limit or 'All UNCERTAIN rows'}")
    print(f"Mode: {'DRY RUN - no Excel updates' if dry_run else 'LIVE - will update Excel'}")
    print()

    start_time = time.time()

    # Find UNCERTAIN rows
    uncertain_rows = find_uncertain_rows(sheet_name, limit)
    if not uncertain_rows:
        print("No UNCERTAIN rows found to process")
        return False

    print(f"Processing {len(uncertain_rows)} UNCERTAIN rows...")
    print()

    # Track results
    upgrade_to_likely = []
    keep_uncertain = []
    no_images_found = []
    processing_errors = []

    # Process each row with enhanced comparison
    for i, row_data in enumerate(uncertain_rows):
        progress = f"[{i+1}/{len(uncertain_rows)}]"
        row_num = row_data['row_num']

        print(f"{progress} Row {row_num}: {row_data['part_num']} vs SKP {row_data['skp_num']}")

        try:
            # Scrape current data
            brand = sheet_name.upper().strip()
            if brand == "FOUR SEASONS ":
                brand = "FOUR SEASONS"

            print(f"  Scraping {brand} {row_data['part_num']}...", end=" ")
            anchor_data = scrape_rockauto(row_data['part_num'], brand)

            print(f"SKP {row_data['skp_num']}...", end=" ")
            skp_data = scrape_rockauto(row_data['skp_num'], "SKP")

            if not anchor_data.get('found'):
                print("No anchor images")
                no_images_found.append(row_data)
                continue

            if not skp_data.get('found'):
                print("No SKP images")
                no_images_found.append(row_data)
                continue

            # Enhanced image comparison
            print("Comparing...")
            comparison = compare_part_images_enhanced(anchor_data, skp_data)
            row_data['enhanced_comparison'] = comparison

            verdict = comparison.get('verdict', 'UNCERTAIN')
            score = comparison.get('similarity_score', 0.0)
            confidence = comparison.get('confidence', 'LOW')

            if verdict == "LIKELY":
                print(f"  -> UPGRADE to LIKELY (score: {score:.2f}, {confidence})")
                upgrade_to_likely.append(row_data)
            else:
                print(f"  -> Keep UNCERTAIN (score: {score:.2f}, {confidence})")
                keep_uncertain.append(row_data)

        except Exception as e:
            print(f"  -> ERROR: {e}")
            row_data['error'] = str(e)
            processing_errors.append(row_data)

        # Minimal pause for system responsiveness
        time.sleep(0.05)

    elapsed = time.time() - start_time

    # Results summary
    print()
    print("=" * 55)
    print("ENHANCED ANALYSIS RESULTS")
    print("=" * 55)
    print(f"Processing time: {elapsed:.1f}s ({elapsed/len(uncertain_rows):.1f}s per row)")
    print(f"Total processed: {len(uncertain_rows)}")
    print(f"")
    print(f"Results breakdown:")
    print(f"  Upgrade to LIKELY: {len(upgrade_to_likely):>3} ({len(upgrade_to_likely)/len(uncertain_rows)*100:4.1f}%)")
    print(f"  Keep UNCERTAIN:    {len(keep_uncertain):>3} ({len(keep_uncertain)/len(uncertain_rows)*100:4.1f}%)")
    print(f"  No images:         {len(no_images_found):>3} ({len(no_images_found)/len(uncertain_rows)*100:4.1f}%)")
    print(f"  Errors:            {len(processing_errors):>3} ({len(processing_errors)/len(uncertain_rows)*100:4.1f}%)")
    print()

    processable_rows = len(uncertain_rows) - len(no_images_found) - len(processing_errors)
    if processable_rows > 0:
        upgrade_rate = len(upgrade_to_likely) / processable_rows * 100
        print(f"Success rate: {upgrade_rate:.1f}% of processable rows upgraded to LIKELY")
        print(f"(Baseline moondream: ~67%, Enhanced target: 74%+)")

        if upgrade_rate >= 74:
            print(f"\\n>> TARGET ACHIEVED! Enhanced system performing at {upgrade_rate:.1f}% vs 67% baseline")
        elif upgrade_rate > 67:
            print(f"\\n>> IMPROVEMENT! Enhanced system at {upgrade_rate:.1f}% vs 67% baseline (+{upgrade_rate-67:.1f}%)")
        else:
            print(f"\\n>> Below baseline. Needs further tuning. Current: {upgrade_rate:.1f}% vs 67% baseline")

    # Apply upgrades to Excel
    if not dry_run and upgrade_to_likely:
        print(f"\\nApplying {len(upgrade_to_likely)} upgrades to Excel...")
        success = apply_upgrades_to_excel(sheet_name, upgrade_to_likely)
        if success:
            print("[OK] Excel updated successfully")
        else:
            print("[ERROR] Failed to update Excel")
            return False
    elif dry_run:
        print("\\n[DRY RUN] Skipping Excel updates")

    print("\\nEnhanced image analysis complete!")
    return True

def apply_upgrades_to_excel(sheet_name, upgrade_rows):
    """Apply UNCERTAIN -> LIKELY upgrades to Excel file"""
    try:
        print(f"Opening Excel file for updates...")
        wb = load_workbook(str(EXCEL_FILE))
        ws = wb[sheet_name]

        # Apply each upgrade
        for row_data in upgrade_rows:
            row_num = row_data['row_num']
            comparison = row_data['enhanced_comparison']

            # Column J: MATCH RESULT
            ws.cell(row=row_num, column=10, value="LIKELY")

            # Column K: CONFIDENCE %
            confidence_pct = int(comparison['similarity_score'] * 100)
            ws.cell(row=row_num, column=11, value=confidence_pct)

            # Column L: MATCH REASON
            ws.cell(row=row_num, column=12, value=comparison['reasoning'][:255])  # Excel limit

            # Column P: LAST CHECKED
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
            ws.cell(row=row_num, column=16, value=timestamp)

            # Apply LIKELY color (light green)
            fill = PatternFill("solid", fgColor="90EE90")
            ws.cell(row=row_num, column=10).fill = fill

        # Crash-safe save
        tmp_file = str(EXCEL_FILE) + ".tmp"
        wb.save(tmp_file)
        os.replace(tmp_file, str(EXCEL_FILE))
        wb.close()

        print(f"Successfully updated {len(upgrade_rows)} rows to LIKELY")
        return True

    except Exception as e:
        print(f"Excel update failed: {e}")
        return False

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Enhanced Image Analysis - Phase 1")
    parser.add_argument("sheet",
                       choices=["GMB", "Four Seasons ", "SMP", "Anchor", "Dorman"],
                       help="Sheet to process")
    parser.add_argument("--limit", type=int, default=5,
                       help="Limit rows for testing (default: 5)")
    parser.add_argument("--dry-run", action="store_true",
                       help="Preview without updating Excel")

    args = parser.parse_args()

    success = run_enhanced_image_analysis(
        args.sheet,
        args.limit,
        args.dry_run
    )

    if not success:
        sys.exit(1)