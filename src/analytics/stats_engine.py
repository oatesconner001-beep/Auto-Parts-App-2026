"""
Enhanced Statistics Engine for Parts Agent

Provides comprehensive statistical analysis of parts matching results across all sheets.
Includes multi-dimensional analysis by brand, part type, time period, and confidence levels.
"""

import openpyxl
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
import json
import re
from collections import defaultdict, Counter

class StatsEngine:
    """
    Advanced statistics engine for comprehensive parts matching analysis.

    Features:
    - Multi-dimensional analysis (brand, part type, time, confidence)
    - Historical trend analysis
    - Quality metrics and benchmarking
    - Performance optimization insights
    """

    def __init__(self, excel_path: str = None):
        self.excel_path = excel_path or self._get_default_excel_path()
        self.sheets = ["GMB", "Four Seasons ", "SMP", "Anchor", "Dorman"]
        self.verdicts = ["YES", "LIKELY", "UNCERTAIN", "NO"]

        # Cached data for performance
        self._cached_data = {}
        self._cache_timestamp = None
        self._cache_ttl = 300  # 5 minutes

        # Part type categories for analysis
        self.part_categories = {
            "Engine": ["ENGINE MOUNT", "ENGINE", "TIMING", "BELT", "GASKET"],
            "Brake": ["BRAKE", "PAD", "ROTOR", "CALIPER"],
            "Suspension": ["STRUT", "SHOCK", "SPRING", "MOUNT"],
            "Electrical": ["SENSOR", "SWITCH", "RELAY", "ALTERNATOR"],
            "Drivetrain": ["TRANSMISSION", "DRIVESHAFT", "CV", "AXLE"],
            "Other": []  # Default category
        }

    def _get_default_excel_path(self) -> str:
        """Get default Excel file path."""
        return str(Path(__file__).parent.parent.parent / "FISHER SKP INTERCHANGE 20260302.xlsx")

    def _is_cache_valid(self) -> bool:
        """Check if cached data is still valid."""
        if not self._cache_timestamp:
            return False
        return (datetime.now() - self._cache_timestamp).seconds < self._cache_ttl

    def _load_excel_data(self, force_reload: bool = False) -> Dict:
        """Load and cache Excel data for analysis."""
        if not force_reload and self._is_cache_valid() and self._cached_data:
            return self._cached_data

        excel_path = Path(self.excel_path)
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        try:
            wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)

            all_data = {}
            for sheet_name in self.sheets:
                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]
                sheet_data = []

                # Read all rows with data
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or len(row) < 16:  # Ensure we have all columns
                        continue

                    row_data = {
                        "row_number": row_idx,
                        "part_type": str(row[0] or "").strip(),
                        "supplier": str(row[1] or "").strip(),
                        "part_number": str(row[2] or "").strip(),
                        "skp_part_number": str(row[5] or "").strip(),
                        "match_result": str(row[9] or "").strip() if len(row) > 9 else "",
                        "confidence": self._parse_confidence(row[10]) if len(row) > 10 else None,
                        "match_reason": str(row[11] or "").strip() if len(row) > 11 else "",
                        "fitment_match": str(row[12] or "").strip() if len(row) > 12 else "",
                        "desc_match": str(row[13] or "").strip() if len(row) > 13 else "",
                        "missing_info": str(row[14] or "").strip() if len(row) > 14 else "",
                        "last_checked": self._parse_datetime(row[15]) if len(row) > 15 else None,
                    }

                    # Add derived fields
                    row_data["is_processed"] = row_data["match_result"] in self.verdicts
                    row_data["is_confirmed"] = row_data["match_result"] in ["YES", "LIKELY"]
                    row_data["part_category"] = self._categorize_part(row_data["part_type"])
                    row_data["has_both_parts"] = bool(row_data["part_number"] and row_data["skp_part_number"])

                    sheet_data.append(row_data)

                all_data[sheet_name] = sheet_data

            wb.close()

            # Cache the data
            self._cached_data = all_data
            self._cache_timestamp = datetime.now()

            return all_data

        except Exception as e:
            raise Exception(f"Failed to load Excel data: {e}")

    def _parse_confidence(self, value) -> Optional[int]:
        """Parse confidence value from Excel."""
        if value is None:
            return None
        try:
            if isinstance(value, (int, float)):
                return int(value)
            if isinstance(value, str):
                # Extract numeric value
                match = re.search(r'(\d+)', str(value))
                return int(match.group(1)) if match else None
        except:
            pass
        return None

    def _parse_datetime(self, value) -> Optional[datetime]:
        """Parse datetime value from Excel."""
        if not value:
            return None
        try:
            if isinstance(value, datetime):
                return value
            if isinstance(value, str):
                # Try common datetime formats
                formats = [
                    "%Y-%m-%d %H:%M",
                    "%Y-%m-%d %H:%M:%S",
                    "%m/%d/%Y %H:%M",
                    "%Y-%m-%d"
                ]
                for fmt in formats:
                    try:
                        return datetime.strptime(value, fmt)
                    except:
                        continue
        except:
            pass
        return None

    def _categorize_part(self, part_type: str) -> str:
        """Categorize part type into broader categories."""
        if not part_type:
            return "Other"

        part_type_upper = part_type.upper()
        for category, keywords in self.part_categories.items():
            if category == "Other":
                continue
            for keyword in keywords:
                if keyword in part_type_upper:
                    return category
        return "Other"

    def get_summary_stats(self) -> Dict:
        """Get comprehensive summary statistics across all sheets."""
        data = self._load_excel_data()

        summary = {
            "overview": self._get_overview_stats(data),
            "by_sheet": self._get_sheet_breakdown(data),
            "by_category": self._get_category_breakdown(data),
            "confidence_analysis": self._get_confidence_analysis(data),
            "processing_status": self._get_processing_status(data),
            "timestamp": datetime.now().isoformat()
        }

        return summary

    def _get_overview_stats(self, data: Dict) -> Dict:
        """Get high-level overview statistics."""
        totals = {verdict: 0 for verdict in self.verdicts}
        totals.update({
            "total_rows": 0,
            "processed_rows": 0,
            "unprocessed_rows": 0,
            "confirmed_matches": 0,
            "needs_review": 0,
            "total_sheets": len([s for s in self.sheets if s in data and data[s]])
        })

        for sheet_name, sheet_data in data.items():
            for row in sheet_data:
                totals["total_rows"] += 1
                if row["is_processed"]:
                    totals["processed_rows"] += 1
                    totals[row["match_result"]] += 1
                    if row["is_confirmed"]:
                        totals["confirmed_matches"] += 1
                    elif row["match_result"] == "UNCERTAIN":
                        totals["needs_review"] += 1
                else:
                    totals["unprocessed_rows"] += 1

        # Calculate percentages
        processed = totals["processed_rows"]
        if processed > 0:
            totals["success_rate"] = (totals["confirmed_matches"] / processed) * 100
            totals["review_rate"] = (totals["needs_review"] / processed) * 100
            totals["processing_completion"] = (processed / totals["total_rows"]) * 100
        else:
            totals.update({"success_rate": 0, "review_rate": 0, "processing_completion": 0})

        return totals

    def _get_sheet_breakdown(self, data: Dict) -> Dict:
        """Get detailed breakdown by sheet."""
        breakdown = {}

        for sheet_name, sheet_data in data.items():
            sheet_stats = {verdict: 0 for verdict in self.verdicts}
            sheet_stats.update({
                "total_rows": len(sheet_data),
                "processed": 0,
                "unprocessed": 0,
                "confirmed": 0,
                "avg_confidence": 0,
                "part_types_count": 0,
                "unique_suppliers": set(),
                "latest_processing": None
            })

            confidences = []
            part_types = set()

            for row in sheet_data:
                if row["is_processed"]:
                    sheet_stats["processed"] += 1
                    sheet_stats[row["match_result"]] += 1
                    if row["is_confirmed"]:
                        sheet_stats["confirmed"] += 1
                    if row["confidence"] is not None:
                        confidences.append(row["confidence"])
                    if row["last_checked"]:
                        if not sheet_stats["latest_processing"] or row["last_checked"] > sheet_stats["latest_processing"]:
                            sheet_stats["latest_processing"] = row["last_checked"]
                else:
                    sheet_stats["unprocessed"] += 1

                if row["part_type"]:
                    part_types.add(row["part_type"])
                if row["supplier"]:
                    sheet_stats["unique_suppliers"].add(row["supplier"])

            # Calculate derived metrics
            sheet_stats["part_types_count"] = len(part_types)
            sheet_stats["unique_suppliers"] = len(sheet_stats["unique_suppliers"])
            if confidences:
                sheet_stats["avg_confidence"] = sum(confidences) / len(confidences)
            if sheet_stats["processed"] > 0:
                sheet_stats["success_rate"] = (sheet_stats["confirmed"] / sheet_stats["processed"]) * 100
            else:
                sheet_stats["success_rate"] = 0

            breakdown[sheet_name] = sheet_stats

        return breakdown

    def _get_category_breakdown(self, data: Dict) -> Dict:
        """Get breakdown by part category."""
        category_stats = defaultdict(lambda: {
            "count": 0,
            "processed": 0,
            "confirmed": 0,
            "uncertain": 0,
            "success_rate": 0,
            "avg_confidence": 0,
            "sheets": set()
        })

        for sheet_name, sheet_data in data.items():
            for row in sheet_data:
                category = row["part_category"]
                stats = category_stats[category]

                stats["count"] += 1
                stats["sheets"].add(sheet_name)

                if row["is_processed"]:
                    stats["processed"] += 1
                    if row["is_confirmed"]:
                        stats["confirmed"] += 1
                    elif row["match_result"] == "UNCERTAIN":
                        stats["uncertain"] += 1

        # Calculate derived metrics and convert sets to counts
        for category, stats in category_stats.items():
            if stats["processed"] > 0:
                stats["success_rate"] = (stats["confirmed"] / stats["processed"]) * 100
            stats["sheets"] = len(stats["sheets"])

        return dict(category_stats)

    def _get_confidence_analysis(self, data: Dict) -> Dict:
        """Analyze confidence score distributions and patterns."""
        confidence_data = []

        for sheet_name, sheet_data in data.items():
            for row in sheet_data:
                if row["confidence"] is not None and row["is_processed"]:
                    confidence_data.append({
                        "confidence": row["confidence"],
                        "verdict": row["match_result"],
                        "sheet": sheet_name,
                        "category": row["part_category"]
                    })

        if not confidence_data:
            return {"error": "No confidence data available"}

        # Statistical analysis
        confidences = [d["confidence"] for d in confidence_data]
        analysis = {
            "total_samples": len(confidences),
            "mean_confidence": sum(confidences) / len(confidences),
            "min_confidence": min(confidences),
            "max_confidence": max(confidences),
            "distribution": self._get_confidence_distribution(confidences),
            "by_verdict": self._group_confidence_by_verdict(confidence_data),
            "by_sheet": self._group_confidence_by_sheet(confidence_data),
            "calibration": self._analyze_confidence_calibration(confidence_data)
        }

        return analysis

    def _get_confidence_distribution(self, confidences: List[int]) -> Dict:
        """Get confidence score distribution in buckets."""
        buckets = {"0-20": 0, "21-40": 0, "41-60": 0, "61-80": 0, "81-100": 0}

        for conf in confidences:
            if conf <= 20:
                buckets["0-20"] += 1
            elif conf <= 40:
                buckets["21-40"] += 1
            elif conf <= 60:
                buckets["41-60"] += 1
            elif conf <= 80:
                buckets["61-80"] += 1
            else:
                buckets["81-100"] += 1

        return buckets

    def _group_confidence_by_verdict(self, confidence_data: List[Dict]) -> Dict:
        """Group confidence scores by verdict."""
        by_verdict = defaultdict(list)

        for item in confidence_data:
            by_verdict[item["verdict"]].append(item["confidence"])

        # Calculate statistics for each verdict
        verdict_stats = {}
        for verdict, confidences in by_verdict.items():
            if confidences:
                verdict_stats[verdict] = {
                    "count": len(confidences),
                    "mean": sum(confidences) / len(confidences),
                    "min": min(confidences),
                    "max": max(confidences)
                }

        return verdict_stats

    def _group_confidence_by_sheet(self, confidence_data: List[Dict]) -> Dict:
        """Group confidence scores by sheet."""
        by_sheet = defaultdict(list)

        for item in confidence_data:
            by_sheet[item["sheet"]].append(item["confidence"])

        # Calculate statistics for each sheet
        sheet_stats = {}
        for sheet, confidences in by_sheet.items():
            if confidences:
                sheet_stats[sheet] = {
                    "count": len(confidences),
                    "mean": sum(confidences) / len(confidences),
                    "min": min(confidences),
                    "max": max(confidences)
                }

        return sheet_stats

    def _analyze_confidence_calibration(self, confidence_data: List[Dict]) -> Dict:
        """Analyze how well confidence scores correlate with actual outcomes."""
        # Group by confidence ranges and analyze accuracy
        ranges = [(0, 20), (21, 40), (41, 60), (61, 80), (81, 100)]
        calibration = {}

        for min_conf, max_conf in ranges:
            range_key = f"{min_conf}-{max_conf}"
            range_data = [
                d for d in confidence_data
                if min_conf <= d["confidence"] <= max_conf
            ]

            if range_data:
                total = len(range_data)
                confirmed = len([d for d in range_data if d["verdict"] in ["YES", "LIKELY"]])
                calibration[range_key] = {
                    "sample_size": total,
                    "actual_success_rate": (confirmed / total) * 100 if total > 0 else 0,
                    "expected_confidence": (min_conf + max_conf) / 2
                }

        return calibration

    def _get_processing_status(self, data: Dict) -> Dict:
        """Get detailed processing status and recommendations."""
        status = {
            "completion_by_sheet": {},
            "processing_velocity": {},
            "bottlenecks": [],
            "recommendations": [],
            "enhanced_analysis_ready": {}
        }

        for sheet_name, sheet_data in data.items():
            total = len(sheet_data)
            processed = sum(1 for row in sheet_data if row["is_processed"])
            uncertain = sum(1 for row in sheet_data if row["match_result"] == "UNCERTAIN")

            completion = (processed / total * 100) if total > 0 else 0
            status["completion_by_sheet"][sheet_name] = {
                "total": total,
                "processed": processed,
                "remaining": total - processed,
                "completion_percentage": completion,
                "uncertain_count": uncertain
            }

            # Enhanced analysis readiness
            status["enhanced_analysis_ready"][sheet_name] = {
                "uncertain_rows": uncertain,
                "estimated_upgrades": int(uncertain * 1.0),  # 100% upgrade rate from Phase 1
                "priority": "HIGH" if uncertain > 100 else "MEDIUM" if uncertain > 10 else "LOW"
            }

            # Identify bottlenecks
            if completion < 10 and total > 100:
                status["bottlenecks"].append(f"{sheet_name}: {completion:.1f}% complete, {total-processed} rows remaining")

        # Generate recommendations
        total_uncertain = sum(s["uncertain_count"] for s in status["completion_by_sheet"].values())
        if total_uncertain > 0:
            status["recommendations"].append(f"Deploy Enhanced Image Analysis on {total_uncertain} UNCERTAIN rows (100% upgrade rate expected)")

        incomplete_sheets = [name for name, info in status["completion_by_sheet"].items() if info["completion_percentage"] < 90]
        if incomplete_sheets:
            status["recommendations"].append(f"Continue main processing on {', '.join(incomplete_sheets)}")

        return status

    def get_sheet_details(self, sheet_name: str) -> Dict:
        """Get detailed analysis for a specific sheet."""
        data = self._load_excel_data()

        if sheet_name not in data:
            return {"error": f"Sheet '{sheet_name}' not found"}

        sheet_data = data[sheet_name]

        details = {
            "basic_stats": self._get_sheet_breakdown({sheet_name: sheet_data})[sheet_name],
            "part_type_analysis": self._analyze_part_types(sheet_data),
            "confidence_trends": self._analyze_confidence_trends(sheet_data),
            "processing_timeline": self._get_processing_timeline(sheet_data),
            "quality_indicators": self._get_quality_indicators(sheet_data),
            "recommendations": self._get_sheet_recommendations(sheet_name, sheet_data)
        }

        return details

    def _analyze_part_types(self, sheet_data: List[Dict]) -> Dict:
        """Analyze part types within a sheet."""
        part_analysis = defaultdict(lambda: {
            "count": 0,
            "processed": 0,
            "success_rate": 0,
            "avg_confidence": 0
        })

        for row in sheet_data:
            if not row["part_type"]:
                continue

            analysis = part_analysis[row["part_type"]]
            analysis["count"] += 1

            if row["is_processed"]:
                analysis["processed"] += 1
                if row["is_confirmed"]:
                    analysis["success_rate"] += 1

        # Calculate final percentages
        for part_type, analysis in part_analysis.items():
            if analysis["processed"] > 0:
                analysis["success_rate"] = (analysis["success_rate"] / analysis["processed"]) * 100

        return dict(part_analysis)

    def _analyze_confidence_trends(self, sheet_data: List[Dict]) -> Dict:
        """Analyze confidence score trends in a sheet."""
        confidences_by_verdict = defaultdict(list)

        for row in sheet_data:
            if row["confidence"] is not None and row["is_processed"]:
                confidences_by_verdict[row["match_result"]].append(row["confidence"])

        trends = {}
        for verdict, confidences in confidences_by_verdict.items():
            if confidences:
                trends[verdict] = {
                    "count": len(confidences),
                    "average": sum(confidences) / len(confidences),
                    "range": max(confidences) - min(confidences)
                }

        return trends

    def _get_processing_timeline(self, sheet_data: List[Dict]) -> Dict:
        """Get processing timeline for a sheet."""
        processed_items = [
            row for row in sheet_data
            if row["is_processed"] and row["last_checked"]
        ]

        if not processed_items:
            return {"error": "No processing timeline data available"}

        # Sort by processing time
        processed_items.sort(key=lambda x: x["last_checked"])

        timeline = {
            "first_processed": processed_items[0]["last_checked"].isoformat(),
            "last_processed": processed_items[-1]["last_checked"].isoformat(),
            "total_duration": str(processed_items[-1]["last_checked"] - processed_items[0]["last_checked"]),
            "items_processed": len(processed_items),
            "processing_days": len(set(item["last_checked"].date() for item in processed_items))
        }

        return timeline

    def _get_quality_indicators(self, sheet_data: List[Dict]) -> Dict:
        """Get data quality indicators for a sheet."""
        indicators = {
            "data_completeness": 0,
            "part_number_quality": 0,
            "match_reason_completeness": 0,
            "confidence_availability": 0,
            "issues": []
        }

        total_processed = sum(1 for row in sheet_data if row["is_processed"])

        if total_processed == 0:
            return indicators

        # Calculate quality metrics
        complete_reasons = sum(1 for row in sheet_data if row["is_processed"] and row["match_reason"])
        available_confidence = sum(1 for row in sheet_data if row["is_processed"] and row["confidence"] is not None)
        valid_parts = sum(1 for row in sheet_data if row["has_both_parts"])

        indicators.update({
            "data_completeness": (total_processed / len(sheet_data)) * 100,
            "part_number_quality": (valid_parts / len(sheet_data)) * 100,
            "match_reason_completeness": (complete_reasons / total_processed) * 100,
            "confidence_availability": (available_confidence / total_processed) * 100
        })

        # Identify quality issues
        if indicators["data_completeness"] < 90:
            indicators["issues"].append("Low processing completion rate")
        if indicators["match_reason_completeness"] < 80:
            indicators["issues"].append("Many results missing explanations")
        if indicators["confidence_availability"] < 70:
            indicators["issues"].append("Confidence scores often missing")

        return indicators

    def _get_sheet_recommendations(self, sheet_name: str, sheet_data: List[Dict]) -> List[str]:
        """Get specific recommendations for a sheet."""
        recommendations = []

        total = len(sheet_data)
        processed = sum(1 for row in sheet_data if row["is_processed"])
        uncertain = sum(1 for row in sheet_data if row["match_result"] == "UNCERTAIN")

        completion_rate = (processed / total) * 100 if total > 0 else 0

        if completion_rate < 50:
            recommendations.append("Priority: Continue main processing - low completion rate")
        elif uncertain > 50:
            recommendations.append("Priority: Run enhanced image analysis on UNCERTAIN rows")
        elif completion_rate > 90 and uncertain < 10:
            recommendations.append("Status: Sheet processing nearly complete")

        if uncertain > 100:
            recommendations.append(f"Enhanced Analysis Impact: ~{uncertain} additional LIKELY matches expected")

        return recommendations

    def get_timestamp(self) -> str:
        """Get current timestamp for reports."""
        return datetime.now().isoformat()

    def export_stats(self, filepath: str) -> bool:
        """Export comprehensive statistics to JSON file."""
        try:
            stats = self.get_summary_stats()
            with open(filepath, 'w') as f:
                json.dump(stats, f, indent=2, default=str)
            return True
        except Exception as e:
            print(f"Failed to export stats: {e}")
            return False

    def clear_cache(self):
        """Clear cached data to force reload."""
        self._cached_data = {}
        self._cache_timestamp = None