"""
Data Quality Analyzer for Parts Agent

Analyzes data quality, consistency, and reliability across all processing results.
Identifies issues and provides recommendations for data improvement.
"""

import re
import openpyxl
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Tuple, Set
from collections import defaultdict, Counter
import statistics
import json

class DataQualityAnalyzer:
    """
    Comprehensive data quality analysis for parts matching results.

    Features:
    - Data completeness validation
    - Consistency checking across sheets
    - Quality scoring and benchmarking
    - Anomaly detection
    - Data reliability assessment
    """

    def __init__(self, excel_path: str = None):
        self.excel_path = excel_path or self._get_default_excel_path()
        self.sheets = ["GMB", "Four Seasons ", "SMP", "Anchor", "Dorman"]
        self.verdicts = ["YES", "LIKELY", "UNCERTAIN", "NO"]

        # Quality benchmarks and thresholds
        self.quality_thresholds = {
            "completeness": {"excellent": 95, "good": 85, "fair": 70, "poor": 50},
            "consistency": {"excellent": 98, "good": 90, "fair": 80, "poor": 70},
            "confidence_reliability": {"excellent": 90, "good": 80, "fair": 70, "poor": 60},
            "oem_match_quality": {"excellent": 85, "good": 75, "fair": 65, "poor": 50}
        }

        # Data validation patterns
        self.validation_patterns = {
            "part_number": r"^[A-Z0-9\-\.]+$",
            "oem_reference": r"^[A-Z0-9\-\.]+$",
            "confidence": r"^\d{1,3}$",
            "part_type": r"^[A-Z\s]+$"
        }

    def _get_default_excel_path(self) -> str:
        """Get default Excel file path."""
        return str(Path(__file__).parent.parent.parent / "FISHER SKP INTERCHANGE 20260302.xlsx")

    def _load_excel_data(self) -> Dict:
        """Load Excel data for quality analysis."""
        excel_path = Path(self.excel_path)
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        try:
            wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
            all_data = {}

            for sheet_name in self.sheets:
                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]
                sheet_data = []

                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or len(row) < 16:
                        continue

                    row_data = {
                        "row_number": row_idx,
                        "part_type": str(row[0] or "").strip(),
                        "supplier": str(row[1] or "").strip(),
                        "part_number": str(row[2] or "").strip(),
                        "skp_part_number": str(row[5] or "").strip(),
                        "match_result": str(row[9] or "").strip() if len(row) > 9 else "",
                        "confidence": self._parse_confidence(row[10]) if len(row) > 10 else None,
                        "match_reason": str(row[11] or "").strip() if len(row) > 11 else "",
                        "fitment_match": str(row[12] or "").strip() if len(row) > 12 else "",
                        "desc_match": str(row[13] or "").strip() if len(row) > 13 else "",
                        "missing_info": str(row[14] or "").strip() if len(row) > 14 else "",
                        "last_checked": self._parse_datetime(row[15]) if len(row) > 15 else None,
                    }

                    sheet_data.append(row_data)

                all_data[sheet_name] = sheet_data

            wb.close()
            return all_data

        except Exception as e:
            raise Exception(f"Failed to load Excel data: {e}")

    def _parse_confidence(self, value) -> Optional[int]:
        """Parse confidence value from Excel."""
        if value is None:
            return None
        try:
            if isinstance(value, (int, float)):
                return int(value)
            if isinstance(value, str):
                match = re.search(r'(\d+)', str(value))
                return int(match.group(1)) if match else None
        except:
            pass
        return None

    def _parse_datetime(self, value) -> Optional[datetime]:
        """Parse datetime value from Excel."""
        if not value:
            return None
        try:
            if isinstance(value, datetime):
                return value
            if isinstance(value, str):
                formats = [
                    "%Y-%m-%d %H:%M",
                    "%Y-%m-%d %H:%M:%S",
                    "%m/%d/%Y %H:%M",
                    "%Y-%m-%d"
                ]
                for fmt in formats:
                    try:
                        return datetime.strptime(value, fmt)
                    except:
                        continue
        except:
            pass
        return None

    def get_quality_summary(self) -> Dict:
        """Get comprehensive data quality summary."""
        data = self._load_excel_data()

        summary = {
            "overall_quality": self._calculate_overall_quality(data),
            "completeness_analysis": self._analyze_completeness(data),
            "consistency_analysis": self._analyze_consistency(data),
            "confidence_reliability": self._analyze_confidence_reliability(data),
            "oem_matching_quality": self._analyze_oem_matching_quality(data),
            "anomaly_detection": self._detect_anomalies(data),
            "data_validation": self._validate_data_formats(data),
            "sheet_comparison": self._compare_sheet_quality(data),
            "recommendations": [],
            "timestamp": datetime.now().isoformat()
        }

        # Generate quality recommendations
        summary["recommendations"] = self._generate_quality_recommendations(summary)

        return summary

    def _calculate_overall_quality(self, data: Dict) -> Dict:
        """Calculate overall data quality score."""
        total_rows = sum(len(sheet_data) for sheet_data in data.values())
        processed_rows = sum(
            len([row for row in sheet_data if row["match_result"] in self.verdicts])
            for sheet_data in data.values()
        )

        # Quality components
        completeness_score = self._calculate_completeness_score(data)
        consistency_score = self._calculate_consistency_score(data)
        accuracy_score = self._calculate_accuracy_score(data)
        reliability_score = self._calculate_reliability_score(data)

        # Weighted overall score
        weights = {"completeness": 0.3, "consistency": 0.2, "accuracy": 0.3, "reliability": 0.2}
        overall_score = (
            completeness_score * weights["completeness"] +
            consistency_score * weights["consistency"] +
            accuracy_score * weights["accuracy"] +
            reliability_score * weights["reliability"]
        )

        # Determine quality grade
        grade = self._score_to_grade(overall_score)

        return {
            "overall_score": overall_score,
            "grade": grade,
            "total_rows": total_rows,
            "processed_rows": processed_rows,
            "processing_completion": (processed_rows / total_rows * 100) if total_rows > 0 else 0,
            "component_scores": {
                "completeness": completeness_score,
                "consistency": consistency_score,
                "accuracy": accuracy_score,
                "reliability": reliability_score
            }
        }

    def _calculate_completeness_score(self, data: Dict) -> float:
        """Calculate data completeness score."""
        total_fields = 0
        complete_fields = 0

        for sheet_data in data.values():
            for row in sheet_data:
                if row["match_result"] in self.verdicts:  # Only count processed rows
                    # Essential fields that should be complete
                    essential_fields = [
                        "part_type", "part_number", "skp_part_number",
                        "match_result", "match_reason"
                    ]

                    for field in essential_fields:
                        total_fields += 1
                        if row.get(field) and row[field].strip():
                            complete_fields += 1

        return (complete_fields / total_fields * 100) if total_fields > 0 else 0

    def _calculate_consistency_score(self, data: Dict) -> float:
        """Calculate data consistency score."""
        consistency_checks = []

        # Check verdict-confidence consistency
        verdict_confidence_consistent = self._check_verdict_confidence_consistency(data)
        consistency_checks.append(verdict_confidence_consistent)

        # Check part number format consistency
        part_format_consistent = self._check_part_format_consistency(data)
        consistency_checks.append(part_format_consistent)

        # Check reason-verdict consistency
        reason_verdict_consistent = self._check_reason_verdict_consistency(data)
        consistency_checks.append(reason_verdict_consistent)

        return statistics.mean(consistency_checks) * 100

    def _calculate_accuracy_score(self, data: Dict) -> float:
        """Calculate estimated accuracy score based on indicators."""
        accuracy_indicators = []

        # High confidence matches should correlate with positive verdicts
        high_conf_accuracy = self._analyze_high_confidence_accuracy(data)
        accuracy_indicators.append(high_conf_accuracy)

        # OEM matches should have higher success rates
        oem_match_accuracy = self._analyze_oem_match_accuracy(data)
        accuracy_indicators.append(oem_match_accuracy)

        # Consistent results across similar part types
        part_type_consistency = self._analyze_part_type_consistency(data)
        accuracy_indicators.append(part_type_consistency)

        return statistics.mean(accuracy_indicators) * 100

    def _calculate_reliability_score(self, data: Dict) -> float:
        """Calculate data reliability score."""
        reliability_factors = []

        # Confidence score distribution should be reasonable
        confidence_distribution = self._analyze_confidence_distribution(data)
        reliability_factors.append(confidence_distribution)

        # Processing timestamps should be recent and consistent
        timestamp_quality = self._analyze_timestamp_quality(data)
        reliability_factors.append(timestamp_quality)

        # Error rates should be low
        error_rate_quality = self._analyze_error_indicators(data)
        reliability_factors.append(error_rate_quality)

        return statistics.mean(reliability_factors) * 100

    def _score_to_grade(self, score: float) -> str:
        """Convert numeric score to letter grade."""
        if score >= 90:
            return "A"
        elif score >= 80:
            return "B"
        elif score >= 70:
            return "C"
        elif score >= 60:
            return "D"
        else:
            return "F"

    def _analyze_completeness(self, data: Dict) -> Dict:
        """Detailed completeness analysis."""
        completeness = {
            "by_field": {},
            "by_sheet": {},
            "critical_gaps": [],
            "completion_trends": {}
        }

        # Analyze by field
        field_stats = defaultdict(lambda: {"total": 0, "complete": 0})

        for sheet_name, sheet_data in data.items():
            sheet_completeness = {"total": 0, "complete": 0}

            for row in sheet_data:
                if row["match_result"] in self.verdicts:
                    sheet_completeness["total"] += 1

                    # Check each important field
                    fields_to_check = {
                        "part_type": row["part_type"],
                        "part_number": row["part_number"],
                        "skp_part_number": row["skp_part_number"],
                        "match_reason": row["match_reason"],
                        "confidence": row["confidence"],
                        "last_checked": row["last_checked"]
                    }

                    row_complete = True
                    for field_name, field_value in fields_to_check.items():
                        field_stats[field_name]["total"] += 1

                        if field_value is not None and str(field_value).strip():
                            field_stats[field_name]["complete"] += 1
                        else:
                            row_complete = False

                    if row_complete:
                        sheet_completeness["complete"] += 1

            # Calculate sheet completeness percentage
            if sheet_completeness["total"] > 0:
                completeness["by_sheet"][sheet_name] = {
                    "completion_rate": (sheet_completeness["complete"] / sheet_completeness["total"]) * 100,
                    "total_processed": sheet_completeness["total"],
                    "fully_complete": sheet_completeness["complete"]
                }

        # Calculate field completion rates
        for field_name, stats in field_stats.items():
            if stats["total"] > 0:
                completion_rate = (stats["complete"] / stats["total"]) * 100
                completeness["by_field"][field_name] = {
                    "completion_rate": completion_rate,
                    "complete_count": stats["complete"],
                    "total_count": stats["total"]
                }

                # Identify critical gaps
                if completion_rate < 80:
                    completeness["critical_gaps"].append({
                        "field": field_name,
                        "completion_rate": completion_rate,
                        "missing_count": stats["total"] - stats["complete"]
                    })

        return completeness

    def _analyze_consistency(self, data: Dict) -> Dict:
        """Detailed consistency analysis."""
        consistency = {
            "verdict_confidence_alignment": self._check_verdict_confidence_consistency(data),
            "part_format_standardization": self._analyze_part_format_standardization(data),
            "cross_sheet_consistency": self._analyze_cross_sheet_consistency(data),
            "reasoning_consistency": self._analyze_reasoning_consistency(data),
            "inconsistencies": []
        }

        # Identify specific inconsistencies
        inconsistencies = []

        # Find high confidence low verdicts
        for sheet_name, sheet_data in data.items():
            for row in sheet_data:
                if row["confidence"] and row["confidence"] > 80 and row["match_result"] in ["NO", "UNCERTAIN"]:
                    inconsistencies.append({
                        "type": "high_confidence_low_verdict",
                        "sheet": sheet_name,
                        "row": row["row_number"],
                        "confidence": row["confidence"],
                        "verdict": row["match_result"]
                    })

        # Find low confidence high verdicts
        for sheet_name, sheet_data in data.items():
            for row in sheet_data:
                if row["confidence"] and row["confidence"] < 60 and row["match_result"] in ["YES", "LIKELY"]:
                    inconsistencies.append({
                        "type": "low_confidence_high_verdict",
                        "sheet": sheet_name,
                        "row": row["row_number"],
                        "confidence": row["confidence"],
                        "verdict": row["match_result"]
                    })

        consistency["inconsistencies"] = inconsistencies[:50]  # Limit to 50 examples

        return consistency

    def _analyze_confidence_reliability(self, data: Dict) -> Dict:
        """Analyze confidence score reliability."""
        confidence_data = []

        for sheet_name, sheet_data in data.items():
            for row in sheet_data:
                if row["confidence"] is not None and row["match_result"] in self.verdicts:
                    confidence_data.append({
                        "confidence": row["confidence"],
                        "verdict": row["match_result"],
                        "sheet": sheet_name
                    })

        if not confidence_data:
            return {"error": "No confidence data available"}

        # Analyze confidence distribution
        confidences = [d["confidence"] for d in confidence_data]
        reliability = {
            "total_samples": len(confidences),
            "mean_confidence": statistics.mean(confidences),
            "median_confidence": statistics.median(confidences),
            "std_deviation": statistics.stdev(confidences) if len(confidences) > 1 else 0,
            "distribution_analysis": self._analyze_confidence_distribution_detailed(confidences),
            "calibration_analysis": self._analyze_confidence_calibration(confidence_data),
            "verdict_confidence_correlation": self._analyze_verdict_confidence_correlation(confidence_data)
        }

        return reliability

    def _analyze_oem_matching_quality(self, data: Dict) -> Dict:
        """Analyze OEM reference matching quality."""
        oem_analysis = {
            "oem_extraction_rate": 0,
            "oem_format_quality": 0,
            "oem_match_correlation": 0,
            "common_oem_patterns": [],
            "quality_issues": []
        }

        total_processed = 0
        oem_mentions = 0
        oem_formats_valid = 0

        # Pattern to find OEM references in reasons
        oem_pattern = r'\b[A-Z0-9]{6,}\b'  # Basic OEM pattern

        for sheet_name, sheet_data in data.items():
            for row in sheet_data:
                if row["match_result"] in self.verdicts:
                    total_processed += 1

                    # Check for OEM references in match reason
                    if row["match_reason"]:
                        oem_refs = re.findall(oem_pattern, row["match_reason"])
                        if oem_refs:
                            oem_mentions += 1

                            # Check format quality
                            valid_format = all(len(ref) >= 6 and ref.isalnum() for ref in oem_refs)
                            if valid_format:
                                oem_formats_valid += 1

        if total_processed > 0:
            oem_analysis["oem_extraction_rate"] = (oem_mentions / total_processed) * 100

        if oem_mentions > 0:
            oem_analysis["oem_format_quality"] = (oem_formats_valid / oem_mentions) * 100

        return oem_analysis

    def _detect_anomalies(self, data: Dict) -> Dict:
        """Detect data anomalies and outliers."""
        anomalies = {
            "statistical_outliers": [],
            "pattern_anomalies": [],
            "temporal_anomalies": [],
            "logical_inconsistencies": []
        }

        # Statistical outliers in confidence scores
        all_confidences = []
        for sheet_data in data.values():
            for row in sheet_data:
                if row["confidence"] is not None:
                    all_confidences.append(row["confidence"])

        if len(all_confidences) > 10:
            mean_conf = statistics.mean(all_confidences)
            std_conf = statistics.stdev(all_confidences)

            for sheet_name, sheet_data in data.items():
                for row in sheet_data:
                    if row["confidence"] is not None:
                        z_score = abs(row["confidence"] - mean_conf) / std_conf
                        if z_score > 3:  # 3 standard deviations
                            anomalies["statistical_outliers"].append({
                                "sheet": sheet_name,
                                "row": row["row_number"],
                                "confidence": row["confidence"],
                                "z_score": z_score,
                                "verdict": row["match_result"]
                            })

        # Pattern anomalies - unusual part numbers
        for sheet_name, sheet_data in data.items():
            for row in sheet_data:
                part_num = row["part_number"]
                if part_num and len(part_num) > 20:  # Unusually long part numbers
                    anomalies["pattern_anomalies"].append({
                        "type": "unusually_long_part_number",
                        "sheet": sheet_name,
                        "row": row["row_number"],
                        "part_number": part_num[:30] + "..." if len(part_num) > 30 else part_num
                    })

        # Temporal anomalies - processing timestamps
        recent_threshold = datetime.now() - timedelta(days=30)
        very_old_threshold = datetime.now() - timedelta(days=365)

        for sheet_name, sheet_data in data.items():
            for row in sheet_data:
                if row["last_checked"]:
                    if row["last_checked"] < very_old_threshold:
                        anomalies["temporal_anomalies"].append({
                            "type": "very_old_processing",
                            "sheet": sheet_name,
                            "row": row["row_number"],
                            "timestamp": row["last_checked"].isoformat()
                        })

        # Logical inconsistencies
        for sheet_name, sheet_data in data.items():
            for row in sheet_data:
                # Empty part numbers but processed
                if not row["part_number"] and row["match_result"] in self.verdicts:
                    anomalies["logical_inconsistencies"].append({
                        "type": "processed_without_part_number",
                        "sheet": sheet_name,
                        "row": row["row_number"]
                    })

                # High confidence NO matches (unusual)
                if row["confidence"] and row["confidence"] > 90 and row["match_result"] == "NO":
                    anomalies["logical_inconsistencies"].append({
                        "type": "high_confidence_no_match",
                        "sheet": sheet_name,
                        "row": row["row_number"],
                        "confidence": row["confidence"]
                    })

        # Limit results to prevent overwhelming output
        for key in anomalies:
            anomalies[key] = anomalies[key][:20]

        return anomalies

    def _validate_data_formats(self, data: Dict) -> Dict:
        """Validate data format compliance."""
        validation = {
            "format_compliance": {},
            "validation_errors": [],
            "format_recommendations": []
        }

        total_validations = 0
        passed_validations = 0

        for sheet_name, sheet_data in data.items():
            for row in sheet_data:
                if row["match_result"] in self.verdicts:
                    # Validate part numbers
                    if row["part_number"]:
                        total_validations += 1
                        if re.match(self.validation_patterns["part_number"], row["part_number"]):
                            passed_validations += 1
                        else:
                            validation["validation_errors"].append({
                                "type": "invalid_part_number_format",
                                "sheet": sheet_name,
                                "row": row["row_number"],
                                "value": row["part_number"][:20]
                            })

                    # Validate confidence scores
                    if row["confidence"] is not None:
                        total_validations += 1
                        if 0 <= row["confidence"] <= 100:
                            passed_validations += 1
                        else:
                            validation["validation_errors"].append({
                                "type": "invalid_confidence_range",
                                "sheet": sheet_name,
                                "row": row["row_number"],
                                "value": row["confidence"]
                            })

        if total_validations > 0:
            validation["format_compliance"]["overall_rate"] = (passed_validations / total_validations) * 100

        # Limit error examples
        validation["validation_errors"] = validation["validation_errors"][:30]

        return validation

    def _compare_sheet_quality(self, data: Dict) -> Dict:
        """Compare data quality across sheets."""
        comparison = {}

        for sheet_name, sheet_data in data.items():
            processed_rows = [row for row in sheet_data if row["match_result"] in self.verdicts]

            if not processed_rows:
                continue

            # Calculate sheet-specific metrics
            sheet_metrics = {
                "total_rows": len(sheet_data),
                "processed_rows": len(processed_rows),
                "completion_rate": (len(processed_rows) / len(sheet_data)) * 100,
                "avg_confidence": 0,
                "success_rate": 0,
                "has_confidence_scores": 0,
                "has_match_reasons": 0,
                "data_freshness": 0
            }

            # Calculate averages
            confidences = [row["confidence"] for row in processed_rows if row["confidence"] is not None]
            if confidences:
                sheet_metrics["avg_confidence"] = statistics.mean(confidences)
                sheet_metrics["has_confidence_scores"] = (len(confidences) / len(processed_rows)) * 100

            successful = len([row for row in processed_rows if row["match_result"] in ["YES", "LIKELY"]])
            sheet_metrics["success_rate"] = (successful / len(processed_rows)) * 100

            reasons = len([row for row in processed_rows if row["match_reason"]])
            sheet_metrics["has_match_reasons"] = (reasons / len(processed_rows)) * 100

            # Data freshness (percentage of rows processed in last 30 days)
            recent_threshold = datetime.now() - timedelta(days=30)
            recent_rows = [row for row in processed_rows if row["last_checked"] and row["last_checked"] > recent_threshold]
            sheet_metrics["data_freshness"] = (len(recent_rows) / len(processed_rows)) * 100

            comparison[sheet_name] = sheet_metrics

        return comparison

    def _generate_quality_recommendations(self, summary: Dict) -> List[str]:
        """Generate data quality improvement recommendations."""
        recommendations = []

        # Overall quality recommendations
        overall_quality = summary.get("overall_quality", {})
        if overall_quality.get("overall_score", 0) < 80:
            recommendations.append("Overall data quality needs improvement - focus on completeness and consistency")

        # Completeness recommendations
        completeness = summary.get("completeness_analysis", {})
        if completeness.get("critical_gaps"):
            critical_fields = [gap["field"] for gap in completeness["critical_gaps"][:3]]
            recommendations.append(f"Critical completeness gaps in: {', '.join(critical_fields)}")

        # Consistency recommendations
        consistency = summary.get("consistency_analysis", {})
        if len(consistency.get("inconsistencies", [])) > 10:
            recommendations.append("High number of consistency issues - review confidence score calibration")

        # Confidence reliability recommendations
        confidence_reliability = summary.get("confidence_reliability", {})
        if confidence_reliability.get("std_deviation", 0) > 25:
            recommendations.append("High confidence score variance - implement confidence calibration")

        # Anomaly recommendations
        anomalies = summary.get("anomaly_detection", {})
        total_anomalies = sum(len(anomalies.get(key, [])) for key in ["statistical_outliers", "pattern_anomalies", "logical_inconsistencies"])
        if total_anomalies > 20:
            recommendations.append("Significant data anomalies detected - implement data validation checks")

        # Sheet comparison recommendations
        sheet_comparison = summary.get("sheet_comparison", {})
        if sheet_comparison:
            lowest_quality_sheet = min(sheet_comparison.items(), key=lambda x: x[1].get("completion_rate", 0))
            if lowest_quality_sheet[1].get("completion_rate", 0) < 50:
                recommendations.append(f"Priority: Improve {lowest_quality_sheet[0]} sheet quality (low completion rate)")

        return recommendations

    # Helper methods for detailed analysis

    def _check_verdict_confidence_consistency(self, data: Dict) -> float:
        """Check consistency between verdict and confidence scores."""
        consistent_count = 0
        total_count = 0

        for sheet_data in data.values():
            for row in sheet_data:
                if row["confidence"] is not None and row["match_result"] in self.verdicts:
                    total_count += 1

                    # Define expected confidence ranges for each verdict
                    expected_ranges = {
                        "YES": (80, 100),
                        "LIKELY": (60, 90),
                        "UNCERTAIN": (30, 70),
                        "NO": (0, 50)
                    }

                    verdict = row["match_result"]
                    confidence = row["confidence"]

                    if verdict in expected_ranges:
                        min_conf, max_conf = expected_ranges[verdict]
                        if min_conf <= confidence <= max_conf:
                            consistent_count += 1

        return (consistent_count / total_count) if total_count > 0 else 0

    def _check_part_format_consistency(self, data: Dict) -> float:
        """Check part number format consistency."""
        consistent_count = 0
        total_count = 0

        for sheet_data in data.values():
            for row in sheet_data:
                if row["part_number"]:
                    total_count += 1
                    if re.match(self.validation_patterns["part_number"], row["part_number"]):
                        consistent_count += 1

        return (consistent_count / total_count) if total_count > 0 else 0

    def _check_reason_verdict_consistency(self, data: Dict) -> float:
        """Check consistency between match reason and verdict."""
        consistent_count = 0
        total_count = 0

        # Keywords that should indicate positive matches
        positive_keywords = ["match", "same", "oem", "identical", "equivalent"]
        negative_keywords = ["no match", "different", "incompatible", "mismatch"]

        for sheet_data in data.values():
            for row in sheet_data:
                if row["match_reason"] and row["match_result"] in self.verdicts:
                    total_count += 1
                    reason_lower = row["match_reason"].lower()

                    # Check if reason aligns with verdict
                    has_positive = any(keyword in reason_lower for keyword in positive_keywords)
                    has_negative = any(keyword in reason_lower for keyword in negative_keywords)

                    verdict = row["match_result"]

                    if verdict in ["YES", "LIKELY"] and has_positive and not has_negative:
                        consistent_count += 1
                    elif verdict == "NO" and has_negative and not has_positive:
                        consistent_count += 1
                    elif verdict == "UNCERTAIN":
                        # UNCERTAIN should have mixed or unclear signals
                        consistent_count += 1

        return (consistent_count / total_count) if total_count > 0 else 0

    def _analyze_confidence_distribution_detailed(self, confidences: List[int]) -> Dict:
        """Detailed confidence score distribution analysis."""
        if not confidences:
            return {"error": "No confidence data"}

        # Create distribution buckets
        buckets = {
            "0-20": 0, "21-40": 0, "41-60": 0, "61-80": 0, "81-100": 0
        }

        for conf in confidences:
            if conf <= 20:
                buckets["0-20"] += 1
            elif conf <= 40:
                buckets["21-40"] += 1
            elif conf <= 60:
                buckets["41-60"] += 1
            elif conf <= 80:
                buckets["61-80"] += 1
            else:
                buckets["81-100"] += 1

        total = len(confidences)
        distribution = {
            "buckets": buckets,
            "percentages": {k: (v / total * 100) for k, v in buckets.items()},
            "skewness": self._calculate_skewness(confidences),
            "balance_score": self._calculate_distribution_balance(buckets)
        }

        return distribution

    def _calculate_skewness(self, values: List[float]) -> str:
        """Calculate distribution skewness."""
        if len(values) < 3:
            return "insufficient_data"

        mean = statistics.mean(values)
        median = statistics.median(values)

        if mean > median + 5:
            return "right_skewed"
        elif mean < median - 5:
            return "left_skewed"
        else:
            return "balanced"

    def _calculate_distribution_balance(self, buckets: Dict) -> float:
        """Calculate how balanced the distribution is."""
        total = sum(buckets.values())
        if total == 0:
            return 0

        # Ideal distribution would have roughly equal representation
        ideal_percentage = 100 / len(buckets)
        actual_percentages = [(count / total * 100) for count in buckets.values()]

        # Calculate deviation from ideal
        deviations = [abs(actual - ideal_percentage) for actual in actual_percentages]
        avg_deviation = statistics.mean(deviations)

        # Convert to balance score (lower deviation = higher balance)
        balance_score = max(0, 100 - avg_deviation * 2)

        return balance_score

    def _analyze_high_confidence_accuracy(self, data: Dict) -> float:
        """Analyze accuracy of high confidence predictions."""
        high_conf_matches = 0
        high_conf_total = 0

        for sheet_data in data.values():
            for row in sheet_data:
                if row["confidence"] and row["confidence"] >= 80:
                    high_conf_total += 1
                    if row["match_result"] in ["YES", "LIKELY"]:
                        high_conf_matches += 1

        return (high_conf_matches / high_conf_total) if high_conf_total > 0 else 0.8

    def _analyze_oem_match_accuracy(self, data: Dict) -> float:
        """Analyze accuracy when OEM references are found."""
        oem_success = 0
        oem_total = 0

        for sheet_data in data.values():
            for row in sheet_data:
                if row["match_reason"] and "oem" in row["match_reason"].lower():
                    oem_total += 1
                    if row["match_result"] in ["YES", "LIKELY"]:
                        oem_success += 1

        return (oem_success / oem_total) if oem_total > 0 else 0.7

    def _analyze_part_type_consistency(self, data: Dict) -> float:
        """Analyze consistency of results within part types."""
        part_type_results = defaultdict(list)

        for sheet_data in data.values():
            for row in sheet_data:
                if row["part_type"] and row["match_result"] in self.verdicts:
                    part_type_results[row["part_type"]].append(row["match_result"])

        consistency_scores = []
        for part_type, results in part_type_results.items():
            if len(results) >= 5:  # Only consider types with enough samples
                success_rate = len([r for r in results if r in ["YES", "LIKELY"]]) / len(results)
                # Consistent results (either consistently high or low) are good
                if success_rate >= 0.7 or success_rate <= 0.3:
                    consistency_scores.append(0.9)
                else:
                    consistency_scores.append(0.6)

        return statistics.mean(consistency_scores) if consistency_scores else 0.7

    def _analyze_confidence_distribution(self, data: Dict) -> float:
        """Analyze if confidence distribution is reasonable."""
        confidences = []
        for sheet_data in data.values():
            for row in sheet_data:
                if row["confidence"] is not None:
                    confidences.append(row["confidence"])

        if not confidences:
            return 0.5

        # Good distribution should have reasonable spread
        std_dev = statistics.stdev(confidences) if len(confidences) > 1 else 0
        mean_conf = statistics.mean(confidences)

        # Ideal: moderate spread around middle values
        if 40 <= mean_conf <= 70 and 15 <= std_dev <= 30:
            return 0.9
        elif 30 <= mean_conf <= 80 and 10 <= std_dev <= 35:
            return 0.7
        else:
            return 0.5

    def _analyze_timestamp_quality(self, data: Dict) -> float:
        """Analyze quality of processing timestamps."""
        recent_count = 0
        total_processed = 0

        recent_threshold = datetime.now() - timedelta(days=30)

        for sheet_data in data.values():
            for row in sheet_data:
                if row["match_result"] in self.verdicts:
                    total_processed += 1
                    if row["last_checked"] and row["last_checked"] > recent_threshold:
                        recent_count += 1

        return (recent_count / total_processed) if total_processed > 0 else 0.5

    def _analyze_error_indicators(self, data: Dict) -> float:
        """Analyze indicators of processing errors."""
        error_indicators = 0
        total_processed = 0

        for sheet_data in data.values():
            for row in sheet_data:
                if row["match_result"] in self.verdicts:
                    total_processed += 1
                    # Look for error indicators
                    if (not row["match_reason"] or
                        row["confidence"] is None or
                        (row["match_result"] == "UNCERTAIN" and not row["missing_info"])):
                        error_indicators += 1

        error_rate = (error_indicators / total_processed) if total_processed > 0 else 0
        return max(0, 1 - error_rate)  # Convert error rate to quality score

    def _analyze_part_format_standardization(self, data: Dict) -> Dict:
        """Analyze standardization of part number formats."""
        format_patterns = defaultdict(int)
        total_parts = 0

        for sheet_data in data.values():
            for row in sheet_data:
                if row["part_number"]:
                    total_parts += 1
                    # Analyze format pattern
                    if re.match(r'^[A-Z]+\d+$', row["part_number"]):
                        format_patterns["letters_then_numbers"] += 1
                    elif re.match(r'^\d+[A-Z]*$', row["part_number"]):
                        format_patterns["numbers_then_letters"] += 1
                    elif re.match(r'^[A-Z0-9\-\.]+$', row["part_number"]):
                        format_patterns["alphanumeric_with_separators"] += 1
                    else:
                        format_patterns["other"] += 1

        return {
            "total_parts": total_parts,
            "patterns": dict(format_patterns),
            "standardization_score": max(format_patterns.values()) / total_parts if total_parts > 0 else 0
        }

    def _analyze_cross_sheet_consistency(self, data: Dict) -> float:
        """Analyze consistency across different sheets."""
        # Compare success rates across sheets
        sheet_success_rates = {}

        for sheet_name, sheet_data in data.items():
            processed = [row for row in sheet_data if row["match_result"] in self.verdicts]
            if processed:
                successful = len([row for row in processed if row["match_result"] in ["YES", "LIKELY"]])
                sheet_success_rates[sheet_name] = successful / len(processed)

        if len(sheet_success_rates) < 2:
            return 0.8

        # Calculate consistency based on variance
        rates = list(sheet_success_rates.values())
        if statistics.stdev(rates) < 0.2:  # Less than 20% variance
            return 0.9
        elif statistics.stdev(rates) < 0.3:
            return 0.7
        else:
            return 0.5

    def _analyze_reasoning_consistency(self, data: Dict) -> float:
        """Analyze consistency of reasoning patterns."""
        reason_verdict_pairs = defaultdict(list)

        for sheet_data in data.values():
            for row in sheet_data:
                if row["match_reason"] and row["match_result"] in self.verdicts:
                    # Extract key words from reason
                    reason_lower = row["match_reason"].lower()
                    if "oem" in reason_lower or "match" in reason_lower:
                        reason_verdict_pairs["positive_indicator"].append(row["match_result"])
                    elif "no match" in reason_lower or "different" in reason_lower:
                        reason_verdict_pairs["negative_indicator"].append(row["match_result"])

        consistency_score = 0.7  # Default moderate consistency

        for indicator, verdicts in reason_verdict_pairs.items():
            if len(verdicts) >= 5:
                if indicator == "positive_indicator":
                    positive_rate = len([v for v in verdicts if v in ["YES", "LIKELY"]]) / len(verdicts)
                    if positive_rate > 0.7:
                        consistency_score = max(consistency_score, 0.9)
                elif indicator == "negative_indicator":
                    negative_rate = len([v for v in verdicts if v == "NO"]) / len(verdicts)
                    if negative_rate > 0.7:
                        consistency_score = max(consistency_score, 0.9)

        return consistency_score

    def _analyze_confidence_calibration(self, confidence_data: List[Dict]) -> Dict:
        """Analyze confidence calibration accuracy."""
        ranges = [(0, 20), (21, 40), (41, 60), (61, 80), (81, 100)]
        calibration = {}

        for min_conf, max_conf in ranges:
            range_key = f"{min_conf}-{max_conf}"
            range_data = [
                d for d in confidence_data
                if min_conf <= d["confidence"] <= max_conf
            ]

            if range_data:
                total = len(range_data)
                confirmed = len([d for d in range_data if d["verdict"] in ["YES", "LIKELY"]])
                calibration[range_key] = {
                    "sample_size": total,
                    "actual_success_rate": (confirmed / total) * 100 if total > 0 else 0,
                    "expected_confidence": (min_conf + max_conf) / 2
                }

        return calibration

    def _analyze_verdict_confidence_correlation(self, confidence_data: List[Dict]) -> Dict:
        """Analyze correlation between verdicts and confidence scores."""
        verdict_confidences = defaultdict(list)

        for item in confidence_data:
            verdict_confidences[item["verdict"]].append(item["confidence"])

        correlations = {}
        for verdict, confidences in verdict_confidences.items():
            if len(confidences) >= 5:
                avg_confidence = statistics.mean(confidences)
                correlations[verdict] = {
                    "count": len(confidences),
                    "avg_confidence": avg_confidence,
                    "min_confidence": min(confidences),
                    "max_confidence": max(confidences),
                    "std_dev": statistics.stdev(confidences) if len(confidences) > 1 else 0
                }

        # Calculate overall correlation quality
        expected_order = ["NO", "UNCERTAIN", "LIKELY", "YES"]
        actual_averages = []

        for verdict in expected_order:
            if verdict in correlations:
                actual_averages.append(correlations[verdict]["avg_confidence"])

        # Check if averages are in ascending order (good correlation)
        correlation_quality = 0.7  # Default
        if len(actual_averages) >= 3:
            ascending = all(actual_averages[i] <= actual_averages[i+1] for i in range(len(actual_averages)-1))
            if ascending:
                correlation_quality = 0.9

        return {
            "by_verdict": correlations,
            "correlation_quality": correlation_quality,
            "expected_order": expected_order
        }

    def export_quality_report(self, filepath: str) -> bool:
        """Export comprehensive quality report."""
        try:
            quality_report = self.get_quality_summary()
            with open(filepath, 'w') as f:
                json.dump(quality_report, f, indent=2, default=str)
            return True
        except Exception as e:
            print(f"Failed to export quality report: {e}")
            return False