"""
PHASE 3: Batch Excel Updater — Apply all image comparison results to Excel at once.

Optimized approach:
- Load comparison results from Phase 2
- Batch update Excel file (one save operation)
- Apply color coding
- Generate summary report
"""

import sys
import json
import openpyxl
from pathlib import Path
from datetime import datetime
from openpyxl.styles import PatternFill

_src = Path(__file__).parent
if str(_src) not in sys.path:
    sys.path.insert(0, str(_src))

EXCEL_FILE = Path(__file__).parent.parent / "FISHER SKP INTERCHANGE 20260302.xlsx"
RESULTS_DIR = Path(__file__).parent.parent / "comparison_results"

FILL_YES       = PatternFill("solid", fgColor="00FF00")
FILL_LIKELY    = PatternFill("solid", fgColor="90EE90")
FILL_UNCERTAIN = PatternFill("solid", fgColor="FFFF00")
FILL_NO        = PatternFill("solid", fgColor="FF0000")

FILL_MAP = {
    "YES": FILL_YES,
    "LIKELY": FILL_LIKELY,
    "UNCERTAIN": FILL_UNCERTAIN,
    "NO": FILL_NO,
}

def batch_update_excel(sheet_name, dry_run=False):
    """Phase 3: Apply all image comparison results to Excel"""

    results_file = RESULTS_DIR / f"{sheet_name.lower()}_comparisons.json"

    if not results_file.exists():
        print(f"[FAIL] Results file not found: {results_file}")
        print("   Run Phase 2 (parallel_image_processor.py) first!")
        return False

    if not EXCEL_FILE.exists():
        print(f"[FAIL] Excel file not found: {EXCEL_FILE}")
        return False

    # Load comparison results
    print(f"Loading results: {results_file}")
    with open(results_file, 'r') as f:
        results = json.load(f)

    upgrades = [r for r in results if r.get('upgrade_to_likely', False)]
    print(f"Found {len(upgrades):,} upgrades out of {len(results):,} comparisons")

    if len(upgrades) == 0:
        print("ℹ️ No upgrades to apply")
        return True

    if dry_run:
        print("🔍 DRY RUN - would make these changes:")
        for result in upgrades[:10]:  # Show first 10
            print(f"   Row {result['row_num']}: {result['part_num']} vs {result['skp_num']} -> LIKELY")
        if len(upgrades) > 10:
            print(f"   ... and {len(upgrades) - 10} more")
        return True

    print(f"\n📝 UPDATING EXCEL: {sheet_name}")

    # Load workbook
    print("Loading workbook...")
    wb = openpyxl.load_workbook(str(EXCEL_FILE))
    ws = wb[sheet_name]

    # Apply updates
    updated_rows = 0
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    for result in upgrades:
        row_num = result['row_num']

        try:
            # Update MATCH RESULT (col J)
            ws.cell(row=row_num, column=10).value = "LIKELY"
            ws.cell(row=row_num, column=10).fill = FILL_LIKELY

            # Update CONFIDENCE % (col K)
            confidence = result.get('ai_confidence', 0)
            if confidence == 0 and result.get('cv_score', 0) > 0:
                confidence = int(result['cv_score'] * 100)

            ws.cell(row=row_num, column=11).value = confidence

            # Update MATCH REASON (col L)
            if result.get('ai_result') == 'HIGH_CV':
                reason = f"High visual similarity (CV: {result.get('cv_score', 0):.3f})"
            else:
                reason = f"AI vision analysis: {result.get('ai_result', 'LIKELY')} ({confidence}%)"

            ws.cell(row=row_num, column=12).value = reason

            # Update other columns
            ws.cell(row=row_num, column=13).value = "YES"  # FITMENT MATCH
            ws.cell(row=row_num, column=14).value = "YES"  # DESC MATCH
            ws.cell(row=row_num, column=15).value = ""     # MISSING INFO (clear)
            ws.cell(row=row_num, column=16).value = timestamp  # LAST CHECKED

            updated_rows += 1

        except Exception as e:
            print(f"⚠️ Error updating row {row_num}: {e}")

    print(f"Updated {updated_rows:,} rows")

    # Create backup
    backup_file = EXCEL_FILE.with_suffix(f".backup_{timestamp.replace(':', '-')}.xlsx")
    print(f"Creating backup: {backup_file}")

    # Save original as backup first
    wb_backup = openpyxl.load_workbook(str(EXCEL_FILE))
    wb_backup.save(str(backup_file))
    wb_backup.close()

    # Save updated file
    print("Saving updated Excel file...")
    wb.save(str(EXCEL_FILE))
    wb.close()

    print(f"[OK] EXCEL UPDATE COMPLETE")
    print(f"   Updated: {updated_rows:,} UNCERTAIN -> LIKELY")
    print(f"   Backup: {backup_file}")

    # Generate summary
    generate_summary_report(sheet_name, results, updated_rows)

    return True

def generate_summary_report(sheet_name, results, updated_rows):
    """Generate a summary report of the optimization"""

    report_file = RESULTS_DIR / f"{sheet_name.lower()}_optimization_report.txt"

    with open(report_file, 'w') as f:
        f.write(f"IMAGE ANALYSIS OPTIMIZATION REPORT - {sheet_name}\n")
        f.write("=" * 60 + "\n\n")

        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Sheet: {sheet_name}\n\n")

        f.write("RESULTS SUMMARY:\n")
        f.write(f"  Total comparisons: {len(results):,}\n")
        f.write(f"  Upgraded to LIKELY: {updated_rows:,}\n")
        f.write(f"  Upgrade rate: {updated_rows/len(results)*100:.1f}%\n\n")

        # Breakdown by method
        high_cv = sum(1 for r in results if r.get('ai_result') == 'HIGH_CV')
        ai_vision = sum(1 for r in results if r.get('upgrade_to_likely', False) and r.get('ai_result') != 'HIGH_CV')

        f.write("UPGRADE METHODS:\n")
        f.write(f"  High CV similarity (≥0.75): {high_cv:,}\n")
        f.write(f"  AI vision analysis: {ai_vision:,}\n\n")

        # Performance stats
        if results:
            avg_cv = sum(r.get('cv_score', 0) for r in results) / len(results)
            f.write("PERFORMANCE:\n")
            f.write(f"  Average CV score: {avg_cv:.3f}\n")

            ai_results = [r for r in results if r.get('ai_result') and r['ai_result'] not in ['HIGH_CV']]
            if ai_results:
                avg_ai_conf = sum(r.get('ai_confidence', 0) for r in ai_results) / len(ai_results)
                f.write(f"  Average AI confidence: {avg_ai_conf:.1f}%\n")

        f.write(f"\nDetailed results saved to: comparison results directory\n")

    print(f"REPORT: Report saved: {report_file}")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("sheet", choices=["Anchor", "Dorman", "GMB", "SMP", "Four Seasons"], help="Sheet to process")
    parser.add_argument("--dry-run", action="store_true", help="Preview changes without applying")

    args = parser.parse_args()

    batch_update_excel(args.sheet, args.dry_run)