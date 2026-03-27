"""
Image comparison module for auto parts matching.

Uses multiple free/local computer vision techniques:
1. Perceptual hashing (fast, detects similar images)
2. Feature matching with ORB (detects same part from different angles)
3. Histogram comparison (detects similar shapes/colors)
4. Optional: Gemini Vision API for complex cases
"""

import cv2
import numpy as np
import requests
from PIL import Image
import imagehash
import tempfile
import os
from typing import Dict, Tuple, Optional
import hashlib

# Install required packages
REQUIRED_PACKAGES = [
    "opencv-python",
    "Pillow",
    "imagehash",
    "numpy"
]

def _download_image(url: str) -> Optional[np.ndarray]:
    """Download and convert image to OpenCV format."""
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()

        # Convert directly from bytes to PIL, then to OpenCV
        from io import BytesIO
        pil_img = Image.open(BytesIO(response.content))

        # Convert to RGB if needed
        if pil_img.mode != 'RGB':
            pil_img = pil_img.convert('RGB')

        # Convert PIL to OpenCV (BGR)
        cv_img = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)
        return cv_img

    except Exception as e:
        print(f"[Image] Failed to download {url}: {e}")
        return None

def _perceptual_hash_similarity(url1: str, url2: str) -> float:
    """Compare images using perceptual hashing (fast, good for identical/similar images)."""
    try:
        # Download images directly to memory
        from io import BytesIO
        resp1 = requests.get(url1, timeout=10)
        resp2 = requests.get(url2, timeout=10)

        resp1.raise_for_status()
        resp2.raise_for_status()

        # Load images directly from bytes
        img1 = Image.open(BytesIO(resp1.content))
        img2 = Image.open(BytesIO(resp2.content))

        # Calculate perceptual hashes
        hash1 = imagehash.phash(img1)
        hash2 = imagehash.phash(img2)

        # Calculate similarity (lower hamming distance = more similar)
        hamming_distance = hash1 - hash2
        similarity = max(0, (64 - hamming_distance) / 64)  # Normalize to 0-1

        return similarity

    except Exception as e:
        print(f"[Image] Perceptual hash failed: {e}")
        return 0.0

def _feature_matching_similarity(img1: np.ndarray, img2: np.ndarray) -> float:
    """Compare using ORB feature matching (good for same part, different angles)."""
    try:
        # Convert to grayscale
        gray1 = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY) if len(img1.shape) == 3 else img1
        gray2 = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY) if len(img2.shape) == 3 else img2

        # Create ORB detector
        orb = cv2.ORB_create(nfeatures=500)

        # Find keypoints and descriptors
        kp1, desc1 = orb.detectAndCompute(gray1, None)
        kp2, desc2 = orb.detectAndCompute(gray2, None)

        if desc1 is None or desc2 is None or len(desc1) < 10 or len(desc2) < 10:
            return 0.0

        # Match features using BFMatcher
        bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)
        matches = bf.match(desc1, desc2)

        if len(matches) < 10:
            return 0.0

        # Sort by distance (lower = better match)
        matches = sorted(matches, key=lambda x: x.distance)

        # Calculate similarity based on good matches
        good_matches = [m for m in matches if m.distance < 50]
        similarity = min(len(good_matches) / min(len(kp1), len(kp2)), 1.0)

        return similarity

    except Exception as e:
        print(f"[Image] Feature matching failed: {e}")
        return 0.0

def _histogram_similarity(img1: np.ndarray, img2: np.ndarray) -> float:
    """Compare color histograms (good for similar-looking parts)."""
    try:
        # Calculate histograms for each channel
        hist1_b = cv2.calcHist([img1], [0], None, [256], [0, 256])
        hist1_g = cv2.calcHist([img1], [1], None, [256], [0, 256])
        hist1_r = cv2.calcHist([img1], [2], None, [256], [0, 256])

        hist2_b = cv2.calcHist([img2], [0], None, [256], [0, 256])
        hist2_g = cv2.calcHist([img2], [1], None, [256], [0, 256])
        hist2_r = cv2.calcHist([img2], [2], None, [256], [0, 256])

        # Compare histograms using correlation
        corr_b = cv2.compareHist(hist1_b, hist2_b, cv2.HISTCMP_CORREL)
        corr_g = cv2.compareHist(hist1_g, hist2_g, cv2.HISTCMP_CORREL)
        corr_r = cv2.compareHist(hist1_r, hist2_r, cv2.HISTCMP_CORREL)

        # Average correlation across channels
        avg_correlation = (corr_b + corr_g + corr_r) / 3
        return max(0, avg_correlation)

    except Exception as e:
        print(f"[Image] Histogram comparison failed: {e}")
        return 0.0

def compare_part_images(anchor_data: dict, skp_data: dict) -> dict:
    """
    Compare images between two parts using multiple computer vision techniques.

    Returns:
        similarity_score: float 0-1 (1 = identical)
        method_scores: dict of individual method scores
        confidence: str ("HIGH", "MEDIUM", "LOW")
        reasoning: str explanation
    """
    result = {
        "similarity_score": 0.0,
        "method_scores": {},
        "confidence": "LOW",
        "reasoning": "No images available for comparison",
        "error": None
    }

    # Get image URLs
    anchor_img_url = anchor_data.get('image_url') or (anchor_data.get('images', []) or [None])[0]
    skp_img_url = skp_data.get('image_url') or (skp_data.get('images', []) or [None])[0]

    if not anchor_img_url or not skp_img_url:
        missing = []
        if not anchor_img_url:
            missing.append(f"{anchor_data.get('brand', 'Unknown')} {anchor_data.get('part_number', '')}")
        if not skp_img_url:
            missing.append(f"{skp_data.get('brand', 'Unknown')} {skp_data.get('part_number', '')}")
        result["reasoning"] = f"Missing images for: {', '.join(missing)}"
        return result

    print(f"[Image] Comparing: {anchor_img_url} vs {skp_img_url}")

    try:
        # Method 1: Perceptual hashing (fastest)
        phash_score = _perceptual_hash_similarity(anchor_img_url, skp_img_url)
        result["method_scores"]["perceptual_hash"] = phash_score

        # Download images for OpenCV methods
        anchor_cv = _download_image(anchor_img_url)
        skp_cv = _download_image(skp_img_url)

        if anchor_cv is not None and skp_cv is not None:
            # Method 2: Feature matching
            feature_score = _feature_matching_similarity(anchor_cv, skp_cv)
            result["method_scores"]["feature_matching"] = feature_score

            # Method 3: Histogram comparison
            histogram_score = _histogram_similarity(anchor_cv, skp_cv)
            result["method_scores"]["histogram"] = histogram_score

            # Weighted combination of scores
            weights = {
                "perceptual_hash": 0.4,    # Good for exact/similar images
                "feature_matching": 0.4,   # Good for same part, different angle
                "histogram": 0.2           # Good for overall color/shape similarity
            }

            weighted_score = (
                phash_score * weights["perceptual_hash"] +
                feature_score * weights["feature_matching"] +
                histogram_score * weights["histogram"]
            )

        else:
            # Only perceptual hash worked
            weighted_score = phash_score
            result["reasoning"] = "Only perceptual hash comparison available (download issues)"

        result["similarity_score"] = weighted_score

        # Determine confidence and reasoning
        if weighted_score >= 0.8:
            result["confidence"] = "HIGH"
            result["reasoning"] = f"Strong visual similarity ({weighted_score:.2f}) - likely same part"
        elif weighted_score >= 0.5:
            result["confidence"] = "MEDIUM"
            result["reasoning"] = f"Moderate visual similarity ({weighted_score:.2f}) - possibly related parts"
        else:
            result["confidence"] = "LOW"
            result["reasoning"] = f"Low visual similarity ({weighted_score:.2f}) - likely different parts"

    except Exception as e:
        result["error"] = str(e)
        result["reasoning"] = f"Image comparison failed: {e}"

    return result

def analyze_uncertain_images(excel_file: str, sheet_name: str) -> dict:
    """
    Analyze all UNCERTAIN results in a sheet using image comparison.
    Returns summary statistics and recommendations for match upgrades.
    """
    from openpyxl import load_workbook
    from scraper_local import scrape_rockauto

    print(f"\\n=== ANALYZING {sheet_name.upper()} UNCERTAIN IMAGES ===")

    wb = load_workbook(excel_file, data_only=True)
    sheet = wb[sheet_name]

    uncertain_rows = []

    # Find all UNCERTAIN results
    for row_num in range(2, sheet.max_row + 1):
        match_result = sheet.cell(row=row_num, column=10).value  # Column J
        if match_result == "UNCERTAIN":
            part_num = sheet.cell(row=row_num, column=3).value    # Column C
            skp_num = sheet.cell(row=row_num, column=6).value     # Column F
            part_type = sheet.cell(row=row_num, column=1).value   # Column A

            if part_num and skp_num:
                uncertain_rows.append({
                    "row_num": row_num,
                    "part_num": str(part_num).strip(),
                    "skp_num": str(skp_num).strip(),
                    "part_type": str(part_type or "").strip()
                })

    wb.close()

    print(f"Found {len(uncertain_rows)} UNCERTAIN results to analyze")

    # Image comparison results
    high_similarity = []
    medium_similarity = []
    low_similarity = []
    no_images = []
    errors = []

    for i, row_data in enumerate(uncertain_rows[:50]):  # Limit to first 50 for testing
        print(f"[{i+1}/{min(50, len(uncertain_rows))}] Analyzing row {row_data['row_num']}...")

        try:
            # Re-scrape to get current image URLs
            brand = sheet_name.upper()
            anchor_data = scrape_rockauto(row_data['part_num'], brand=brand)
            skp_data = scrape_rockauto(row_data['skp_num'], brand="SKP")

            if not anchor_data.get('found') or not skp_data.get('found'):
                no_images.append(row_data)
                continue

            # Compare images
            comparison = compare_part_images(anchor_data, skp_data)
            row_data['image_comparison'] = comparison

            confidence = comparison['confidence']
            if confidence == "HIGH":
                high_similarity.append(row_data)
            elif confidence == "MEDIUM":
                medium_similarity.append(row_data)
            elif confidence == "LOW":
                low_similarity.append(row_data)
            else:
                no_images.append(row_data)

        except Exception as e:
            row_data['error'] = str(e)
            errors.append(row_data)
            print(f"  Error: {e}")

    # Summary
    summary = {
        "total_analyzed": min(50, len(uncertain_rows)),
        "high_similarity": len(high_similarity),
        "medium_similarity": len(medium_similarity),
        "low_similarity": len(low_similarity),
        "no_images": len(no_images),
        "errors": len(errors),
        "recommendations": {
            "upgrade_to_likely": high_similarity,
            "needs_manual_review": medium_similarity,
            "likely_no_match": low_similarity
        }
    }

    print(f"\\nSUMMARY:")
    print(f"  High similarity (upgrade to LIKELY): {len(high_similarity)}")
    print(f"  Medium similarity (manual review): {len(medium_similarity)}")
    print(f"  Low similarity (likely NO): {len(low_similarity)}")
    print(f"  No images available: {len(no_images)}")
    print(f"  Errors: {len(errors)}")

    return summary

if __name__ == "__main__":
    # Test with a single part comparison
    print("Testing image comparison...")

    # You can test with known part numbers here
    test_anchor = {"image_url": "https://www.rockauto.com/info/28/3217-000__ra_m.jpg"}
    test_skp = {"image_url": "https://www.rockauto.com/info/983/SKM3217_1___ra_m.jpg"}

    result = compare_part_images(test_anchor, test_skp)
    print(f"Test result: {result}")